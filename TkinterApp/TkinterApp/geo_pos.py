#b8960f47-87d6-49d0-af63-938389333d68
# -*- coding: utf-8 -*-
import enum
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import pandas as pd
import numpy as np
from datetime import datetime, date, time as dtTime, timedelta
import re
import openpyxl as ox
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

import requests
import json
from typing import Optional, Tuple, Dict, Any
from time import sleep as nap
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_coordinates_from_yandex(
    address: str,
    api_key: str,
    max_retries: int = 3,
    timeout: int = 10
) -> Optional[Tuple[float, float]]:
    """
    Получает координаты адреса через Яндекс API.
    Оптимизирована для Москвы и Московской области.
    
    Args:
        address: Строка с адресом (может содержать ошибки)
        api_key: Яндекс API ключ
        max_retries: Максимальное количество попыток
        timeout: Таймаут запроса в секундах
        
    Returns:
        Tuple[широта, долгота] или None если не найдено
    """
    
    def _clean_address(addr: str) -> str:
        """Очистка и нормализация адреса"""
        if not isinstance(addr, str):
            return ""
        
        # Базовые замены и нормализация
        replacements = {
            # Московские сокращения
            r'\bмск\b': 'Москва',
            r'\bг\.москва\b': 'Москва',
            r'\bг москва\b': 'Москва',
            r'\bмос\.обл\b': 'Московская область',
            r'\bмо\b': 'Московская область',
            
            # Опечатки в "Москва"
            r'\bмоскыв\b': 'Москва',
            r'\bмасква\b': 'Москва',
            r'\bмосвка\b': 'Москва',
            r'\bмоскав\b': 'Москва',
            
            # Опечатки в "Московская"
            r'\bмасковская\b': 'Московская',
            r'\bмосковска\b': 'Московская',
            r'\bмосковской\b': 'Московская',
            
            # Стандартизация улиц
            r'\булица\b': 'ул.',
            r'\bпроспект\b': 'пр-т',
            r'\bпр-кт\b': 'пр-т',
            r'\bпрт\.\b': 'пр-т',
            r'\bпроезд\b': 'пр.',
            r'\bпереулок\b': 'пер.',
            r'\bшоссе\b': 'ш.',
            r'\bбульвар\b': 'б-р',
            
            # Стандартизация номеров
            r'\bдом\b': 'д.',
            r'\bкорпус\b': 'к.',
            r'\bстроение\b': 'стр.',
            r'\bквартира\b': 'кв.',
            
            # Убираем лишние пробелы и запятые
            r'\.\s+\.': '.',
            r'\s{2,}': ' ',
            r',\s*,': ',',
        }
        
        cleaned = addr.strip()
        
        # Применяем замены
        for pattern, replacement in replacements.items():
            cleaned = re.sub(pattern, replacement, cleaned, flags=re.IGNORECASE)
        
        # Автоматически добавляем Москву/МО если не указано
        if not re.search(r'(москва|московская|мск|мо)', cleaned, re.IGNORECASE):
            # Проверяем по характерным московским топонимам
            moscow_keywords = [
                r'арбат', r'тверская', r'новый\s+арбат', r'китай-город',
                r'покровка', r'маяковская', r'красная\s+площадь',
                r'кремль', r'метро\s+[а-яё]+', r'цао', r'сао', r'свао',
                r'вао', r'ювао', r'юао', r'юзао', r'зао', r'сзао',
                r'зелао', r'тинао'
            ]
            
            # Проверяем по городам МО
            mo_cities = [
                'балашиха', 'химки', 'подольск', 'королёв', 'мытищи',
                'люберцы', 'красногорск', 'электросталь', 'одинцово',
                'домодедово', 'щёлково', 'раменское', 'серпухов',
                'долгопрудный', 'реутов', 'жуковский', 'лобня', 'дубна'
            ]
            
            address_lower = cleaned.lower()
            
            # Проверяем московские ключевые слова
            is_moscow = any(re.search(keyword, address_lower) for keyword in moscow_keywords)
            
            # Проверяем города МО
            is_mo = any(city in address_lower for city in mo_cities)
            
            if is_moscow:
                cleaned = f"{cleaned}, Москва"
            elif is_mo:
                cleaned = f"{cleaned}, Московская область"
        
        return cleaned
    
    def _is_valid_coords(lat: float, lon: float) -> bool:
        """Проверка, что координаты в районе Москвы/МО"""
        # Московский регион: широта ~55-56, долгота ~37-38
        # Расширенные границы для покрытия всей области
        moscow_region_bounds = {
            'lat_min': 54.0,   # южнее Серпухова
            'lat_max': 57.5,   # севернее Дубны
            'lon_min': 35.0,   # западнее Можайска
            'lon_max': 40.0    # восточнее Егорьевска
        }
        
        return (
            moscow_region_bounds['lat_min'] <= lat <= moscow_region_bounds['lat_max'] and
            moscow_region_bounds['lon_min'] <= lon <= moscow_region_bounds['lon_max']
        )
    
    def _make_yandex_request(addr: str, attempt: int = 1) -> Optional[Dict[str, Any]]:
        """Выполнение запроса к Яндекс API"""
        try:
            url = "https://geocode-maps.yandex.ru/1.x/"
            params = {
                'apikey': api_key,
                'geocode': addr,
                'format': 'json',
                'results': 5,  # Берем несколько результатов для выбора лучшего
                'lang': 'ru_RU',
                'bbox': '35.0,54.0~40.0,57.5'  # Ограничиваем поиск Москвой и областью
            }
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            response = requests.get(
                url,
                params=params,
                headers=headers,
                timeout=timeout
            )
            
            if response.status_code == 200:
                return response.json()
            elif response.status_code == 403:
                logger.error("Ошибка авторизации. Проверьте API ключ.")
                return None
            elif response.status_code == 429:
                if attempt < max_retries:
                    sleep_time = 2 ** attempt  # Экспоненциальная задержка
                    logger.warning(f"Превышен лимит запросов. Ждем {sleep_time} сек...")
                    nap(sleep_time)
                    return _make_yandex_request(addr, attempt + 1)
                else:
                    logger.error("Достигнут лимит повторных попыток")
                    return None
            else:
                logger.error(f"Ошибка API: {response.status_code}")
                return None
                
        except requests.exceptions.Timeout:
            logger.warning(f"Таймаут запроса (попытка {attempt})")
            if attempt < max_retries:
                nap(1)
                return _make_yandex_request(addr, attempt + 1)
        except Exception as e:
            logger.error(f"Ошибка запроса: {e}")
            
        return None
    
    def _extract_best_result(data: Dict[str, Any], original_addr: str) -> Optional[Tuple[float, float]]:
        """Извлечение лучшего результата из ответа API"""
        try:
            members = data.get('response', {}).get('GeoObjectCollection', {}).get('featureMember', [])
            
            if not members:
                return None
            
            # Варианты для выбора лучшего результата
            candidates = []
            
            for i, member in enumerate(members[:5]):  # Ограничиваем первыми 5 результатами
                geo_object = member.get('GeoObject', {})
                name = geo_object.get('name', '')
                description = geo_object.get('description', '')
                
                # Извлекаем координаты
                pos_str = geo_object.get('Point', {}).get('pos', '')
                if not pos_str:
                    continue
                    
                lon_str, lat_str = pos_str.split()
                lat = float(lat_str)
                lon = float(lon_str)
                
                # Проверяем, что в московском регионе
                if not _is_valid_coords(lat, lon):
                    continue
                
                # Вычисляем "качество" результата
                score = 0
                
                # Бонус за точное совпадение Москвы/МО
                original_lower = original_addr.lower()
                result_full = f"{name}, {description}".lower()
                
                if 'москва' in original_lower and 'москва' in result_full:
                    score += 10
                if 'московская' in original_lower and 'московская' in result_full:
                    score += 10
                
                # Бонус за дом (если был указан в запросе)
                if re.search(r'д\.\s*\d+', original_addr, re.IGNORECASE):
                    if 'дом' in result_full or 'д.' in result_full:
                        score += 5
                
                # Штраф за дальность от центра Москвы (для приоритета ближайших)
                moscow_center = (55.7558, 37.6176)
                distance = ((lat - moscow_center[0])**2 + (lon - moscow_center[1])**2)**0.5
                score -= distance * 10
                
                # Бонус за первый результат (обычно самый релевантный)
                score += (5 - i) * 2
                
                candidates.append({
                    'coords': (lat, lon),
                    'score': score,
                    'name': name,
                    'description': description
                })
            
            if not candidates:
                return None
            
            # Выбираем результат с наивысшим score
            best = max(candidates, key=lambda x: x['score'])
            
            logger.info(f"Выбран результат: {best['name']} ({best['description']})")
            return best['coords']
            
        except Exception as e:
            logger.error(f"Ошибка обработки результата: {e}")
            return None
    
    # Основная логика функции
    try:
        if not api_key or api_key.strip() == "":
            logger.error("Не указан API ключ")
            return None
        
        if not address or not isinstance(address, str) or address.strip() == "":
            logger.error("Пустой адрес")
            return None
        
        # Очищаем адрес
        cleaned_address = _clean_address(address)
        logger.info(f"Очищенный адрес: {cleaned_address}")
        
        # Делаем запрос к API
        data = _make_yandex_request(cleaned_address)
        
        if not data:
            # Пробуем без очистки (на случай если очистка испортила)
            logger.info("Пробуем исходный адрес...")
            data = _make_yandex_request(address)
            
        if not data:
            return None
        
        # Извлекаем координаты
        coords = _extract_best_result(data, address)
        latit=f'{coords[0]:.6f}'.replace(',','.')
        longit=f'{coords[1]:.6f}'.replace(',','.')
        if coords:
            logger.info(f"Найдены координаты: {latit}, {longit}")
            return latit, longit
        else:
            logger.warning("Координаты не найдены в московском регионе")
            return None
            
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}")
        return None

# Функция для пакетной обработки адресов
def batch_geocode_addresses(
    addresses: list,
    api_key: str,
    delay: float = 0.1
) -> Dict[str, Optional[Tuple[float, float]]]:
    results = {}
    
    for i, addr in enumerate(addresses):
        print(f"Обработка {i+1}/{len(addresses)}...")
        
        coords = get_coordinates_from_yandex(addr, api_key)
        results[addr] = coords
        results[addr] = coords
        
        if i < len(addresses) - 1:  # Не ждем после последнего
            nap(delay)
    
    return results

def rusLowerCaseWeekDays(num):
    days = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
    return days[num] if 0 <= num <= 6 else ''


def dateConvertion(someDate):
    """Возвращает datetime или None"""
    if isinstance(someDate, datetime):
        return someDate
    if isinstance(someDate, str):
        cleaned = re.sub(r'\s*[\.г\s]*$', '', someDate).strip()
        if cleaned:
            normalized = re.sub(r'[.\-]', '/', cleaned)
            try:
                return datetime.strptime(normalized, '%d/%m/%Y')
            except ValueError:
                return None
        else:
            return None
    return None


def timeCheck(value, fileName):
    if pd.isna(value) or value == '' or value is None:
        return ''
    if isinstance(value, str):
        val = value.strip()
        if not val:
            return ''
        if ':' not in val and '.' in val:
            parts = val.split('.')
            if len(parts) in (2, 3) and all(p.isdigit() for p in parts):
                val = ':'.join(parts)
        try:
            clean_val = re.sub(r'[^\d:]', '', val.split()[0]) if ' ' in val else val
            clean_val = clean_val.replace('AM', '').replace('PM', '').replace('am', '').replace('pm', '').strip()
            if ':' in clean_val:
                parts = clean_val.split(':')
                hour = int(parts[0])
                minute = int(parts[1]) if len(parts) > 1 else 0
                second = int(float(parts[2])) if len(parts) > 2 else 0
                if not (0 <= hour <= 23) or not (0 <= minute <= 59):
                    return ''
                return f"{hour:02d}:{minute:02d}:{second:02d}"
            elif val.isdigit():
                hour = int(val)
                if 0 <= hour <= 23:
                    return f"{hour:02d}:00:00"
            return ''
        except:
            return ''
    elif isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    elif isinstance(value, dtTime):
        dt = datetime.combine(datetime.min, value)
        return dt.strftime("%H:%M:%S")
    elif isinstance(value, (timedelta, pd.Timedelta)):
        total_seconds = int(round(value.total_seconds()))
        hours = total_seconds // 3600 % 24
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    else:
        return timeCheck(str(value), fileName)


def pddChapter(text):
    if pd.isna(text) or not isinstance(text, str):
        return ''
    match = re.search(r'\b(\d+(?:\s*[,\.]\s*\d+)+)', text)
    if match:
        normalized = re.sub(r'\s*,\s*', '.', match.group(1))
        normalized = re.sub(r'\s*\.\s*', '.', normalized)
        return normalized
    return ''


class App:
    def __init__(self, root):
        self.root = root
        root.title('Справки в список')
        root.geometry('500x550')

        self.inputFolder = tk.StringVar()
        self.journal = tk.StringVar()
        self.delayFile = tk.StringVar()

        tk.Label(root, text='Расположение папки со справками:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.inputFolder, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text='Выберите папку', command=self.select_inputFolder).pack(padx=10, pady=5)

        tk.Label(root, text='Отрпавленный журнал:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.journal, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text="Выберите файл", command=self.select_journal).pack(padx=10, pady=5)

        tk.Button(root, text='Запустить обработку', command=self.run_processing, bg='#4CAF50', fg='white',font=('Segoe UI', 10, 'bold')).pack(padx=10, pady=15)

        self.log_text = scrolledtext.ScrolledText(root, wrap=tk.WORD, height=12, font=('Segoe UI', 9))
        self.log_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def select_inputFolder(self):
        folder = filedialog.askdirectory(title='Выберите папку со справками')
        if folder:
            self.inputFolder.set(folder)

    def select_journal(self):
        file = filedialog.askopenfilename(
            title='Выберите файл журнала',
            filetypes=[('Excel files', '*.xlsx')]
        )
        if file:
            self.journal.set(file)


    def run_processing(self):
        if not self.inputFolder.get():
            messagebox.showerror('Ошибка', 'Выберите папку со справками!')
            return
        if not self.journal.get():
            messagebox.showerror('Ошибка', 'Выберите файл журнала!')
            return

        startingTime = datetime.now()
        self.log_text.delete(1.0, tk.END)
        self.log(f'=== Начало обработки: {startingTime.strftime("%H:%M:%S")} ===')

        try:
            # 1. Обработка справок → обновлённый журнал
            self.process_files()

            endingTime = datetime.now()
            timeDif = endingTime - startingTime
            self.log(f'\n=== Готово! Всего: {timeDif.total_seconds():.1f} сек ===')
            messagebox.showinfo('Успех', 'Обработка завершена!\nРезультаты сохранены.')

        except Exception as e:
            self.log(f'\nКРИТИЧЕСКАЯ ОШИБКА: {e}')
            messagebox.showerror('Ошибка', f'Обработка прервана:\n{str(e)}')

    def process_files(self):
        folderPath = self.inputFolder.get()
        journalPath = self.journal.get()

        items = [f for f in os.listdir(folderPath) if not f.startswith(('Журнал', 'Задержк'))]
        total_files = len(items)

        if total_files == 0:
            self.log("Нет файлов для обработки (файлы 'Журнал','Задержки' игнорируются).")
            return

        Result = pd.DataFrame()
        failed_files = []
        empty_fields_warnings = []

        for idx, item in enumerate(items, start=1):
            filePath = os.path.join(folderPath, item)
            self.log(f"[{idx}/{total_files}] Обработка: {item}")

            try:
                if not filePath.endswith('.xls'):
                    try:
                        dataFile = pd.read_excel(filePath, engine='openpyxl').fillna('')
                    except Exception:
                        dataFile = pd.read_excel(filePath, engine='xlrd').fillna('')
                else:
                    dataFile = pd.read_excel(filePath, engine='xlrd').fillna('')

                # --- Дата и день недели ---
                original_date_cell = dataFile.iloc[1][2]
                converted_date = None
                weekDay = ''
                if original_date_cell != '' and not pd.isna(original_date_cell):
                    try:
                        converted_date = dateConvertion(original_date_cell)
                        if converted_date:
                            weekDay = rusLowerCaseWeekDays(converted_date.weekday())
                    except Exception:
                        weekDay = ''

                # --- ФИО водителя ---
                p = str(dataFile.iloc[18][6]).split(' ')
                shortName = f"{p[0]} {p[1][0].upper()+'.' if len(p) > 1 and p[1] else ''}{p[2][0].upper()+'.' if len(p) > 2 and p[2] else ''}".strip()

                # --- Пункт правил ---
                original_pdd_text = dataFile.iloc[40][1]
                pdd_value = ''
                if isinstance(original_pdd_text, str):
                    textLower = original_pdd_text.lower().strip()
                    hasPDD = 'пдд' in textLower
                    hasP = re.search(r'\bп\.?\s*\d', textLower) is not None
                    if hasPDD or hasP:
                        pdd_value = pddChapter(original_pdd_text)

                # --- Стаж ---
                raw_stazh = str(dataFile.iloc[24][9]).strip()
                raw_stazh_park = str(dataFile.iloc[24][20]).strip()
                def process_stazh(raw):
                    if not raw or raw.lower() in ('', 'nan'):
                        return ''
                    if re.fullmatch(r'[\d\s.,]+', raw) and re.search(r'\d', raw):
                        num_part = re.search(r'\d+', raw)
                        if num_part:
                            return int(num_part.group())
                    return raw

                stazh_obshch = process_stazh(raw_stazh)
                stazh_v_parke = process_stazh(raw_stazh_park)

                # --- Пострадавшие ---
                postradavshie_raw = dataFile.iloc[14][8]
                postradavshie = 0
                if postradavshie_raw != '' and str(postradavshie_raw).lower() not in ('нет', ''):
                    try:
                        postradavshie = int(postradavshie_raw)
                    except:
                        postradavshie = 0

                # --- Координаты (оставляем пустыми) ---
                API_KEY = 'b8960f47-87d6-49d0-af63-938389333d68'
                lat, lon = '', ''
                #self.log(f"\nАдрес: {dataFile.iloc[9][6]}")
                coords = get_coordinates_from_yandex(dataFile.iloc[9][6], API_KEY)
                if coords:
                    lat, lon = coords
                    #self.log(f"Координаты: {lat:.6f}, {lon:.6f}")
                    nap(0.4)

                # --- Филиал + Э ---

                filName =str(dataFile.iloc[17][4]).replace('Филиал ', 'Ф').replace('Филилал ', 'Ф').replace('Южный', 'Ю').replace('Северный', 'С').replace('Юго-', 'Ю').replace('Северо-', 'С').replace('Восточный', 'В').replace('Западный', 'З')
                if (str(dataFile.iloc[31][18]).strip()!='' and str(dataFile.iloc[31][18]).strip().startswith('Э')) \
                    or (str(dataFile.iloc[31][18]).strip()=='' and str(dataFile.iloc[33][20]).strip().startswith('4')):
                    filName = filName + '(Э)'
                
                    
                # --- DataFrame ---
                dfData = pd.DataFrame({
                    'Место': '1' if 'европротокол' in str(dataFile.iloc[66][7]).lower() else '',
                    'Дата ДТП': converted_date,
                    'Время ДТП': timeCheck(dataFile.iloc[1][17], item),
                    'День недели': weekDay,
                    ' Место ДТП (Адрес)': dataFile.iloc[9][6],
                    'Район': dataFile.iloc[10][9],
                    'Округ': dataFile.iloc[10][2],
                    'Координаты места ДТП (широта)': lat,
                    'Координаты места ДТП (долгота)': lon,
                    '3-я сторона': '',
                    'Название стороней организации': dataFile.iloc[74][6],
                    'Государственный регистрационный знак сторонего транспорта': dataFile.iloc[73][6],
                    'МГТ': '',
                    'Перевозчик': 'ГУП "Мосгортранс"',
                    'Название филиала': filName,
                    'Название площадки': dataFile.iloc[17][15],
                    'Маршрут': dataFile.iloc[31][8],
                    'Марка автобуса / электробуса': dataFile.iloc[32][8],
                    'Гаражный номер': '' if dataFile.iloc[33][20] == '' else int(dataFile.iloc[33][20]),
                    'Регистрационный номер': dataFile.iloc[33][8],
                    'Водитель': '',
                    'Табельный номер водителя': dataFile.iloc[19][6],
                    'ФИО водителя': shortName,
                    'Гражданство': dataFile.iloc[22][6],
                    'Возраст': dataFile.iloc[21][6],
                    'Стаж общий ': stazh_obshch,
                    'Стаж в парке': stazh_v_parke,
                    'ДТП': '',
                    'Вид ДТП': dataFile.iloc[7][3],
                    'Причина ДТП': dataFile.iloc[40][1],
                    'Виновник ДТП': str(dataFile.iloc[65][9])
                        .replace('Не вина', '3-е лицо')
                        .replace('Вина', 'Перевозчик')
                        .replace('В расследовании', 'Проводится разбор'),
                    'Пункт правил': pdd_value,
                    'Скорость по гланассу КМ/Ч': dataFile.iloc[36][16],
                    'Выделенная полоса (ДА; НЕТ,)': str(dataFile.iloc[12][9]).lower(),
                    'Пострадавшие': '',
                    'Кол-во пострадавших': postradavshie,
                    'в т.ч.    лёгкий   вред здоровью': postradavshie,
                    'в т.ч. средний вред здоровью': '0',
                    'в т.ч. тяжёлый вред здоровью': '0',
                    'Кол-во погибших': '0',
                    'ГК': '0',
                    'Ответст-сть': '',
                    'Постановление': dataFile.iloc[66][7],
                    'Дата постановления': (dateConvertion(dataFile.iloc[67][10])).strftime('%d.%m.%Y') if dateConvertion(dataFile.iloc[67][10]) else '',
                    'Наказание водителя': 'выговор' if dataFile.iloc[65][9] == 'Вина' else '',
                    'Проишествия': '',
                    'Резонансные проишествия': '',
                    'Проишествия с водителями': '',
                    'Проишествия с контролёрами': '',
                    'Проишествия с пассажирами': '',
                    'Кол-во задержек в движении': '',
                    'Сработка АНТИСОН': '',
                    'Кол-во отстранённых водителей.': '',
                    'Проишествия3': ''
                }, index=[0])

                # --- Проверка потерь данных ---
                fields_to_check = {
                    'Дата ДТП': original_date_cell,
                    'Время ДТП': dataFile.iloc[1][17],
                    'ФИО водителя': dataFile.iloc[18][6],
                    'Пункт правил': original_pdd_text,
                    'Стаж общий ': raw_stazh,
                    'Стаж в парке': raw_stazh_park,
                    'Кол-во пострадавших': postradavshie_raw,
                }

                for col_name, orig_val in fields_to_check.items():
                    final_val = dfData[col_name].iloc[0]
                    if (
                        not pd.isna(orig_val) and str(orig_val).strip() != '' and
                        (pd.isna(final_val) or final_val == '')
                    ):
                        empty_fields_warnings.append(f"{item} -> {col_name}: '{orig_val}' -> ''")

                Result = pd.concat([Result, dfData], ignore_index=True)

            except Exception as e:
                failed_files.append(item)
                self.log(f"[{idx}/{total_files}] Ошибка: {e}")

        # --- УДАЛЕНИЕ ДУБЛИКАТОВ ---
        if Result.empty:
            self.log("Нет данных для сохранения.")
            return

        key_cols = ['Гаражный номер', 'Дата ДТП', 'Время ДТП', 'Маршрут']

        for col in key_cols:
            if col not in Result.columns:
                Result[col] = ''

        # 1. Удалить дубликаты внутри новых записей
        Result_before = len(Result)
        Result = Result.drop_duplicates(subset=key_cols, keep='first')
        removed_internal = Result_before - len(Result)
        if removed_internal > 0:
            self.log(f"Удалено {removed_internal} дубликатов внутри новых записей.")

        # 2. Найти добавленные ранее и изменить стаж
        try:
            # Загружаем журнал с помощью openpyxl, чтобы сохранить стили
            wb = ox.load_workbook(journalPath)
            ws = wb['Лист1']

            # Определяем индексы столбцов (1-based)
            headers = [cell.value for cell in ws[1]]
            col_indices = {}
            target_columns = ['Стаж общий ', 'Дата ДТП', 'Время ДТП', 'ФИО водителя', 'Гаражный номер', 'Название филиала', 'Координаты места ДТП (широта)','Координаты места ДТП (долгота)']
            for col_name in target_columns:
                if col_name in headers:
                    col_indices[col_name] = headers.index(col_name) + 1
                else:
                    self.log(f"Предупреждение: столбец '{col_name}' не найден в заголовке журнала.")
                    col_indices = {}
                    break

            if col_indices:
                # Преобразуем Result в список словарей
                result_records = Result.to_dict('records')
                updated_count = 0

                # Обходим строки Excel начиная с 5701 (включительно)
                for row in range(5751, ws.max_row + 1):
                    # Извлекаем значения из Excel
                    date_val = ws.cell(row=row, column=col_indices['Дата ДТП']).value
                    time_val = ws.cell(row=row, column=col_indices['Время ДТП']).value
                    fio_val = ws.cell(row=row, column=col_indices['ФИО водителя']).value
                    garage_val = ws.cell(row=row, column=col_indices['Гаражный номер']).value


                    # Нормализуем для сравнения
                    def norm_date(d):
                        if pd.isna(d) or d == '' or d is None:
                            return ''
                        if isinstance(d, datetime):
                            return d.strftime('%Y-%m-%d')
                        parsed = dateConvertion(str(d))
                        return parsed.strftime('%Y-%m-%d') if parsed else ''

                    def norm_str(s):
                        return str(s).strip() if pd.notna(s) and s != '' and s is not None else ''

                    key_excel = (
                        norm_date(date_val),
                        norm_str(time_val),
                        norm_str(fio_val),
                        norm_str(garage_val)
                    )

                    # Поиск совпадения в новых данных
                    for rec in result_records:
                        key_new = (
                            norm_date(rec['Дата ДТП']),
                            norm_str(rec['Время ДТП']),
                            norm_str(rec['ФИО водителя']),
                            norm_str(rec['Гаражный номер'])
                        )
                        if key_excel == key_new:
                            #new_fil = rec['Название филиала'] 
                            #ws.cell(row=row, column=col_indices['Название филиала'], value=new_fil)                           
                            
                            #new_stazh = rec['Стаж общий ']
                            #ws.cell(row=row, column=col_indices['Стаж общий '], value=new_stazh)

                            #new_stazh = rec['Стаж общий ']
                            #ws.cell(row=row, column=col_indices['Стаж общий '], value=new_stazh)

                            new_lat = rec['Координаты места ДТП (широта)']
                            ws.cell(row=row, column=col_indices['Координаты места ДТП (широта)'], value=new_lat)


                            new_lon = rec['Координаты места ДТП (долгота)']
                            ws.cell(row=row, column=col_indices['Координаты места ДТП (долгота)'], value=new_lon)

                            updated_count += 1
                            break  # достаточно одного совпадения

                self.log(f"Обновлено 'Стаж общий' в {updated_count} строках (начиная с 5701).")

            # Сохраняем обновлённый файл
            output_path = journalPath[:-5] + '_newChanged.xlsx'
            wb.save(output_path)
            self.log(f"\nРезультат сохранён с сохранением форматирования: {output_path}")

        except Exception as e:
            self.log(f"Ошибка при обновлении журнала: {e}")
            # Резерв: сохранить только новые данные без стилей
            output_path = journalPath[:-5] + '_newChanged.xlsx'
            Result.to_excel(output_path, index=False)
            self.log(f"Резервное сохранение (без форматов): {output_path}")

        # --- Итоговый отчёт ---
        if failed_files:
            self.log("\nНеобработанные файлы:")
            for f in failed_files:
                self.log(f"  - {f}")
        else:
            self.log("\nВсе файлы обработаны.")

        if empty_fields_warnings:
            self.log(f"\nПотери данных ({len(empty_fields_warnings)} записей):")
            for w in empty_fields_warnings[:50]:
                self.log(f"  - {w}")
            if len(empty_fields_warnings) > 50:
                self.log(f"  ... и ещё {len(empty_fields_warnings) - 50}")
        else:
            self.log("\nНет потерь данных.")


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()