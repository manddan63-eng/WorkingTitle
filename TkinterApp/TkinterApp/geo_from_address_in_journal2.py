# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from datetime import datetime
import re
import openpyxl as ox
import requests
import csv
import os
from typing import Optional, Tuple
from time import sleep as nap
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_house_number(address: str) -> str:
    if not isinstance(address, str):
        return ""
    address = address.strip()
    patterns = [
        r'(?:д\.?\s*|дом\s*)?(\d+[а-яё]?(?:/\d+)?(?:к\d+)?(?:\s*стр\.?\s*\d+)?)',
        r'(\d+[а-яё]?(?:/\d+)?(?:к\d+)?)\b'
    ]
    for pattern in patterns:
        match = re.search(pattern, address, re.IGNORECASE)
        if match:
            return match.group(1).lower()
    return ""

def extract_street_part(address: str) -> str:
    """Простая попытка выделить улицу до номера дома."""
    if not isinstance(address, str):
        return ""
    # Убираем всё после первого числа (где обычно начинается номер)
    parts = re.split(r'\d', address, maxsplit=1)
    street = parts[0].strip()
    # Убираем общие префиксы
    street = re.sub(r'^(г\.?|город|москва|московская\s+область|ул\.?|улица)\s*', '', street, flags=re.IGNORECASE)
    return street.strip()


def normalize_address(name):
    if not isinstance(name, str):
        return set()
    # Убираем всё лишнее
    cleaned = re.sub(r'[^\w\sа-яё\-]', ' ', name.lower())
    # Убираем общие слова
    cleaned = re.sub(r'\b(улица|ул\.?|проспект|пр-т|пр\.?|бульвар|б-р|шоссе|ш\.?|проезд|переулок|набережная|пл\.)\b', ' ', cleaned)
    # Разбиваем на слова, убираем короткие и пустые
    words = [w.strip() for w in cleaned.split() if len(w) > 1]
    return set(words)

def streets_match(original: str, yandex: str) -> bool:
    """Гибкое сравнение улиц: все ключевые слова из original должны быть в yandex."""
    orig_words = normalize_street_name(original)
    yand_words = normalize_street_name(yandex)

    if not orig_words:
        return True  # не можем проверить — пропускаем

    # Все слова из оригинала должны быть в ответе Яндекса
    return orig_words.issubset(yand_words)


def streets_match(original: str, yandex: str) -> bool:
    if not isinstance(original, str) or not isinstance(yandex, str):
        return True
    orig_low = original.lower()
    yand_low = yandex.lower()

    conflict_rules = [
        ("соломатина", "воскресенские ворота"),
        ("соломатина", "воскресенский"),
        ("воскресенские ворота", "соломатина"),
    ]
    for keyword_orig, forbidden_yandex in conflict_rules:
        if keyword_orig in orig_low and forbidden_yandex in yand_low:
            return False

    if "соломатина" in orig_low and "соломатина" not in yand_low:
        return False

    new_streets = ["соломатина", "космонавта волкова", "героя труда"]
    for street in new_streets:
        if street in orig_low and street not in yand_low:
            return False

    return True

def get_coordinates_from_yandex(address: str, api_key: str, max_retries: int = 3, timeout: int = 5) -> Optional[Tuple[str, str, str]]:
    def _clean_address(addr: str) -> str:
        if not isinstance(addr, str):
            return ""
        replacements = {
            r'\bмск\b': 'Москва',
            r'\bг\.москва\b': 'Москва',
            r'\bг москва\b': 'Москва',
            r'\bмос\.обл\b': 'Московская область',
            r'\bмо\b': 'Московская область',
            r'\bмоскыв\b': 'Москва',
            r'\bмасква\b': 'Москва',
            r'\bмосвка\b': 'Москва',
            r'\bмоскав\b': 'Москва',
            r'\bмасковская\b': 'Московская',
            r'\bмосковска\b': 'Московская',
            r'\bмосковской\b': 'Московская',
            r'\булица\b': 'ул.',
            r'\bпроспект\b': 'пр-т',
            r'\bпр-кт\b': 'пр-т',
            r'\bпрт\.\b': 'пр-т',
            r'\bпроезд\b': 'пр.',
            r'\bпереулок\b': 'пер.',
            r'\bшоссе\b': 'ш.',
            r'\bбульвар\b': 'б-р',
            r'\bдом\b': 'д.',
            r'\bкорпус\b': 'к.',
            r'\bстроение\b': 'стр.',
            r'\bквартира\b': 'кв.',
            r'\.\s+\.': '.',
            r'\s{2,}': ' ',
            r',\s*,': ',',
        }
        cleaned = addr.strip()
        for pattern, replacement in replacements.items():
            cleaned = re.sub(pattern, replacement, cleaned, flags=re.IGNORECASE)

        cleaned = re.sub(
            r'(?:ул\.?\s*)?(?:героя\s+россии\s+)?соломатина',
            'ул. Героя России Соломатина',
            cleaned,
            flags=re.IGNORECASE
        )

        if not re.search(r'(москва|московская|мск|мо)', cleaned, re.IGNORECASE):
            moscow_keywords = [r'арбат', r'тверская', r'новый\s+арбат', r'китай-город',
                               r'покровка', r'маяковская', r'красная\s+площадь', r'кремль',
                               r'метро\s+[а-яё]+', r'цао', r'сао', r'свао', r'вао', r'ювао',
                               r'юао', r'юзао', r'зао', r'сзао', r'зелао', r'тинао']
            mo_cities = ['балашиха', 'химки', 'подольск', 'королёв', 'мытищи', 'люберцы',
                         'красногорск', 'электросталь', 'одинцово', 'домодедово', 'щёлково',
                         'раменское', 'серпухов', 'долгопрудный', 'реутов', 'жуковский', 'лобня', 'дубна']
            addr_lower = cleaned.lower()
            is_moscow = any(re.search(kw, addr_lower) for kw in moscow_keywords)
            is_mo = any(city in addr_lower for city in mo_cities)
            if is_moscow:
                cleaned = f"{cleaned}, Москва"
            elif is_mo:
                cleaned = f"{cleaned}, Московская область"
        return cleaned

    def _is_valid_coords(lat: float, lon: float) -> bool:
        return 54.0 <= lat <= 57.5 and 35.0 <= lon <= 40.0

    def _make_yandex_request(addr: str, attempt: int = 1) -> Optional[dict]:
        try:
            url = "https://geocode-maps.yandex.ru/1.x/"
            params = {
                'apikey': api_key,
                'geocode': addr,
                'format': 'json',
                'results': 1,
                'kind': 'house',
                'lang': 'ru_RU',
                'bbox': '35.0,54.0~40.0,57.5'
            }
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(url, params=params, headers=headers, timeout=timeout)
            if response.status_code == 200:
                return response.json()
            elif response.status_code == 429 and attempt < max_retries:
                nap(10)
                return _make_yandex_request(addr, attempt + 1)
            return None
        except Exception:
            if attempt < max_retries:
                nap(1)
                return _make_yandex_request(addr, attempt + 1)
            return None

    def _extract_best_result( dict, original_addr: str) -> Optional[Tuple[str, float, float]]:
        members = data.get('response', {}).get('GeoObjectCollection', {}).get('featureMember', [])
        if not members:
            return None
        member = members[0]
        geo = member.get('GeoObject', {})
        pos = geo.get('Point', {}).get('pos')
        if not pos:
            return None
        lon_str, lat_str = pos.split()
        lat, lon = float(lat_str), float(lon_str)
        if not _is_valid_coords(lat, lon):
            return None
        full_addr = geo.get('metaDataProperty', {}).get('GeocoderMetaData', {}).get('text', '')
        if not full_addr:
            full_addr = f"{geo.get('name', '')}, {geo.get('description', '')}"
        return full_addr, lat, lon

    if not api_key or not address or not isinstance(address, str) or not address.strip():
        return None

    cleaned = _clean_address(address)
    data = _make_yandex_request(cleaned) or _make_yandex_request(address)
    if not data:
        return None

    result = _extract_best_result(data, address)
    if not result:
        return None

    yandex_addr, lat_f, lon_f = result
    lat_str = f"{lat_f:.6f}".replace(',', '.')
    lon_str = f"{lon_f:.6f}".replace(',', '.')
    return yandex_addr, lat_str, lon_str


class App:
    def __init__(self, root):
        self.root = root
        root.title('Проставка координат + отчёт')
        root.geometry('380x320')

        self.journal = tk.StringVar()

        tk.Label(root, text='Журнал:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.journal, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text="Выберите файл", command=self.select_journal).pack(padx=10, pady=5)

        tk.Button(root, text='Запустить обработку', command=self.run_processing, bg='#4CAF50', fg='white',
                  font=('Segoe UI', 10, 'bold')).pack(padx=10, pady=15)

        self.log_text = scrolledtext.ScrolledText(root, wrap=tk.WORD, height=12, font=('Segoe UI', 9))
        self.log_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def select_journal(self):
        file = filedialog.askopenfilename(title='Выберите файл журнала', filetypes=[('Excel files', '*.xlsx')])
        if file:
            self.journal.set(file)

    def run_processing(self):
        if not self.journal.get():
            messagebox.showerror('Ошибка', 'Выберите файл журнала!')
            return

        startingTime = datetime.now()
        self.log_text.delete(1.0, tk.END)
        self.log(f'=== Начало: {startingTime.strftime("%H:%M:%S")} ===')

        try:
            self.process_files()
            endingTime = datetime.now()
            timeDif = endingTime - startingTime
            self.log(f'\n=== Готово! Всего: {timeDif.total_seconds():.1f} сек ===')
            messagebox.showinfo('Успех', 'Обработка завершена!\nРезультаты и отчёт сохранены.')
        except Exception as e:
            self.log(f'\nКРИТИЧЕСКАЯ ОШИБКА: {e}')
            messagebox.showerror('Ошибка', f'Обработка прервана:\n{str(e)}')

    def process_files(self):
        journalPath = self.journal.get()
        API_KEY = 'b8960f47-87d6-49d0-af63-938389333d68'

        # Имя файла отчёта
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = os.path.join(os.path.dirname(journalPath), f"geocoding_report_{timestamp}.csv")

        try:
            wb = ox.load_workbook(journalPath)
            ws = wb['Лист1']

            headers = [cell.value for cell in ws[1]]
            col_indices = {}
            target_columns = [' Место ДТП (Адрес)', 'Координаты места ДТП (долгота)', 'Координаты места ДТП (широта)']
            for col_name in target_columns:
                if col_name in headers:
                    col_indices[col_name] = headers.index(col_name) + 1
                else:
                    self.log(f"❌ Столбец '{col_name}' не найден!")
                    return

            total = ws.max_row - 5750
            self.log(f'Адресов для обработки: {total}')

            # Открываем CSV-отчёт
            with open(report_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = [
                    'Исходный адрес',
                    'Отправлено в API',
                    'Адрес от Яндекса',
                    'Дом (оригинал)',
                    'Дом (Яндекс)',
                    'Улица (оригинал)',
                    'Улица (Яндекс)',
                    'Решение',
                    'Широта',
                    'Долгота'
                ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
                writer.writeheader()

                for row in range(5751, ws.max_row + 1):
                    addr_val = ws.cell(row=row, column=col_indices[' Место ДТП (Адрес)']).value
                    lat_existing = ws.cell(row=row, column=col_indices['Координаты места ДТП (широта)']).value

                    if not addr_val or not isinstance(addr_val, str) or not addr_val.strip():
                        continue
                    if lat_existing and str(lat_existing).strip() != "":
                        continue

                    # Нормализуем для отправки
                    def _clean_for_log(addr):
                        if not isinstance(addr, str): return ""
                        replacements = {r'\bмск\b': 'Москва', r'\bул\.?\s*': 'ул. '}
                        cleaned = addr.strip()
                        for p, r in replacements.items():
                            cleaned = re.sub(p, r, cleaned, flags=re.IGNORECASE)
                        return cleaned

                    addr_sent = _clean_for_log(addr_val)
                    result = get_coordinates_from_yandex(addr_sent, API_KEY)

                    house_orig = extract_house_number(addr_val)
                    street_orig = extract_street_part(addr_val)

                    if not result:
                        writer.writerow({
                            'Исходный адрес': addr_val,
                            'Отправлено в API': addr_sent,
                            'Адрес от Яндекса': '',
                            'Дом (оригинал)': house_orig,
                            'Дом (Яндекс)': '',
                            'Улица (оригинал)': street_orig,
                            'Улица (Яндекс)': '',
                            'Решение': 'ОШИБКА_API',
                            'Широта': '',
                            'Долгота': ''
                        })
                        self.log(f"❌ Нет результата: {addr_val}")
                        continue

                    yandex_addr, lat, lon = result
                    house_yand = extract_house_number(yandex_addr)
                    street_yand = extract_street_part(yandex_addr)

                    house_ok = not house_orig or not house_yand or (house_orig == house_yand)
                    street_ok = streets_match(addr_val, yandex_addr)
                    accepted = house_ok and street_ok

                    writer.writerow({
                        'Исходный адрес': addr_val,
                        'Отправлено в API': addr_sent,
                        'Адрес от Яндекса': yandex_addr,
                        'Дом (оригинал)': house_orig,
                        'Дом (Яндекс)': house_yand,
                        'Улица (оригинал)': street_orig,
                        'Улица (Яндекс)': street_yand,
                        'Решение': 'ПРИНЯТО' if accepted else 'ОТКЛОНЕНО',
                        'Широта': lat if accepted else '',
                        'Долгота': lon if accepted else ''
                    })

                    if accepted:
                        ws.cell(row=row, column=col_indices['Координаты места ДТП (широта)'], value=lat)
                        ws.cell(row=row, column=col_indices['Координаты места ДТП (долгота)'], value=lon)
                        self.log(f"✅ Принято: {addr_val}")
                    else:
                        self.log(f"⚠️ Отклонено: {addr_val}")

                    nap(0.4)

            output_path = journalPath[:-5] + '_newChanged.xlsx'
            wb.save(output_path)
            self.log(f"\n✅ Результат: {output_path}")
            self.log(f"📄 Отчёт: {report_path}")

        except Exception as e:
            self.log(f"Ошибка: {e}")
            raise


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()