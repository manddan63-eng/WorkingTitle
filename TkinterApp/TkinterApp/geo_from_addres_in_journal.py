import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from datetime import datetime
import re
import openpyxl as ox
import requests
from typing import Optional, Tuple, Dict, Any
from time import sleep as nap
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_coordinates_from_yandex(address: str,api_key: str,max_retries: int = 3,timeout: int = 5) -> Optional[Tuple[float, float]]:

    def _clean_address(addr: str) -> str:
        if not isinstance(addr, str):
            return ""
        
        # Базовые замены и нормализация
        replacements = {
            # Московские сокращения
            r'\bмск\b': 'Москва',
            r'\bг\.москва\b': 'Москва',
            r'\bг москва\b': 'Москва',
            r'\bмос\.обл\b': 'Московская область',
            r'\bмо\b': 'Московская область',
            
            # Опечатки в "Москва"
            r'\bмоскыв\b': 'Москва',
            r'\bмасква\b': 'Москва',
            r'\bмосвка\b': 'Москва',
            r'\bмоскав\b': 'Москва',
            
            # Опечатки в "Московская"
            r'\bмасковская\b': 'Московская',
            r'\bмосковска\b': 'Московская',
            r'\bмосковской\b': 'Московская',
            
            # Стандартизация улиц
            r'\булица\b': 'ул.',
            r'\bпроспект\b': 'пр-т',
            r'\bпр-кт\b': 'пр-т',
            r'\bпрт\.\b': 'пр-т',
            r'\bпроезд\b': 'пр.',
            r'\bпереулок\b': 'пер.',
            r'\bшоссе\b': 'ш.',
            r'\bбульвар\b': 'б-р',
            
            # Стандартизация номеров
            r'\bдом\b': 'д.',
            r'\bкорпус\b': 'к.',
            r'\bстроение\b': 'стр.',
            r'\bквартира\b': 'кв.',
            
            # Убираем лишние пробелы и запятые
            r'\.\s+\.': '.',
            r'\s{2,}': ' ',
            r',\s*,': ',',
        }
        
        cleaned = addr.strip()
        
        # Применяем замены
        for pattern, replacement in replacements.items():
            cleaned = re.sub(pattern, replacement, cleaned, flags=re.IGNORECASE)
        
        # Автоматически добавляем Москву/МО если не указано
        if not re.search(r'(москва|московская|мск|мо)', cleaned, re.IGNORECASE):
            # Проверяем по характерным московским топонимам
            moscow_keywords = [
                r'арбат', r'тверская', r'новый\s+арбат', r'китай-город',
                r'покровка', r'маяковская', r'красная\s+площадь',
                r'кремль', r'метро\s+[а-яё]+', r'цао', r'сао', r'свао',
                r'вао', r'ювао', r'юао', r'юзао', r'зао', r'сзао',
                r'зелао', r'тинао'
            ]
            
            # Проверяем по городам МО
            mo_cities = [
                'балашиха', 'химки', 'подольск', 'королёв', 'мытищи',
                'люберцы', 'красногорск', 'электросталь', 'одинцово',
                'домодедово', 'щёлково', 'раменское', 'серпухов',
                'долгопрудный', 'реутов', 'жуковский', 'лобня', 'дубна'
            ]
            
            address_lower = cleaned.lower()
            
            # Проверяем московские ключевые слова
            is_moscow = any(re.search(keyword, address_lower) for keyword in moscow_keywords)
            
            # Проверяем города МО
            is_mo = any(city in address_lower for city in mo_cities)
            
            if is_moscow:
                cleaned = f"{cleaned}, Москва"
            elif is_mo:
                cleaned = f"{cleaned}, Московская область"
        
        return cleaned
    
    def _is_valid_coords(lat: float, lon: float) -> bool:
        """Проверка, что координаты в районе Москвы/МО"""
        # Московский регион: широта ~55-56, долгота ~37-38
        # Расширенные границы для покрытия всей области
        moscow_region_bounds = {
            'lat_min': 54.0,   # южнее Серпухова
            'lat_max': 57.5,   # севернее Дубны
            'lon_min': 35.0,   # западнее Можайска
            'lon_max': 40.0    # восточнее Егорьевска
        }
        
        return (
            moscow_region_bounds['lat_min'] <= lat <= moscow_region_bounds['lat_max'] and
            moscow_region_bounds['lon_min'] <= lon <= moscow_region_bounds['lon_max']
        )
    
    def _make_yandex_request(addr: str, attempt: int = 1) -> Optional[Dict[str, Any]]:
        """Выполнение запроса к Яндекс API"""
        try:
            url = "https://geocode-maps.yandex.ru/1.x/"
            params = {
                'apikey': api_key,
                'geocode': addr,
                'format': 'json',
                'results': 5,  # Берем несколько результатов для выбора лучшего
                'lang': 'ru_RU',
                'bbox': '35.0,54.0~40.0,57.5'  # Ограничиваем поиск Москвой и областью
            }
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            response = requests.get(
                url,
                params=params,
                headers=headers,
                timeout=timeout
            )
            
            if response.status_code == 200:
                return response.json()
            elif response.status_code == 403:
                #logger.error("Ошибка авторизации. Проверьте API ключ.")
                return None
            elif response.status_code == 429:
                if attempt < max_retries:
                    sleep_time = 10 # ** attempt  # Экспоненциальная задержка
                    logger.warning(f"Превышен лимит запросов. Ждем {sleep_time} сек...")
                    #nap(sleep_time)
                    return _make_yandex_request(addr, attempt + 1)
                else:
                    logger.error("Достигнут лимит повторных попыток")
                    return None
            else:
                #logger.error(f"Ошибка API: {response.status_code}")
                return None
                
        except requests.exceptions.Timeout:
            #logger.warning(f"Таймаут запроса (попытка {attempt})")
            if attempt < max_retries:
                nap(1)
                return _make_yandex_request(addr, attempt + 1)
        except Exception as e:
            logger.error(f"Ошибка запроса: {e}")
            
        return None
    
    def _extract_best_result(data: Dict[str, Any], original_addr: str) -> Optional[Tuple[float, float]]:
        try:
            members = data.get('response', {}).get('GeoObjectCollection', {}).get('featureMember', [])
            
            if not members:
                return None
            
            # Варианты для выбора лучшего результата
            candidates = []
            
            for i, member in enumerate(members[:5]):  # Ограничиваем первыми 5 результатами
                geo_object = member.get('GeoObject', {})
                name = geo_object.get('name', '')
                description = geo_object.get('description', '')
                
                # Извлекаем координаты
                pos_str = geo_object.get('Point', {}).get('pos', '')
                if not pos_str:
                    continue
                    
                lon_str, lat_str = pos_str.split()
                lat = float(lat_str)
                lon = float(lon_str)
                
                # Проверяем, что в московском регионе
                if not _is_valid_coords(lat, lon):
                    continue
                
                # Вычисляем "качество" результата
                score = 0
                
                # Бонус за точное совпадение Москвы/МО
                original_lower = original_addr.lower()
                result_full = f"{name}, {description}".lower()
                
                if 'москва' in original_lower and 'москва' in result_full:
                    score += 10
                if 'московская' in original_lower and 'московская' in result_full:
                    score += 10
                
                # Бонус за дом (если был указан в запросе)
                if re.search(r'д\.\s*\d+', original_addr, re.IGNORECASE):
                    if 'дом' in result_full or 'д.' in result_full:
                        score += 5
                
                # Штраф за дальность от центра Москвы (для приоритета ближайших)
                moscow_center = (55.7558, 37.6176)
                distance = ((lat - moscow_center[0])**2 + (lon - moscow_center[1])**2)**0.5
                score -= distance * 10
                
                # Бонус за первый результат (обычно самый релевантный)
                score += (5 - i) * 2
                
                candidates.append({
                    'coords': (lat, lon),
                    'score': score,
                    'name': name,
                    'description': description
                })
            
            if not candidates:
                return None
            
            # Выбираем результат с наивысшим score
            best = max(candidates, key=lambda x: x['score'])
            
            #logger.info(f"Выбран результат: {best['name']} ({best['description']})")
            return best['coords']
            
        except Exception as e:
            #logger.error(f"Ошибка обработки результата: {e}")
            return None
    
    try:
        if not api_key or api_key.strip() == "":
            logger.error("Не указан API ключ")
            return None
        
        if not address or not isinstance(address, str) or address.strip() == "":
            #logger.error("Пустой адрес")
            return None
        
        # Очищаем адрес
        cleaned_address = _clean_address(address)
        #logger.info(f"Очищенный адрес: {cleaned_address} из {address}")
        
        # Делаем запрос к API
        data = _make_yandex_request(cleaned_address)
        
        if not data:
            # Пробуем без очистки (на случай если очистка испортила)
            #logger.info("Пробуем исходный адрес...")
            data = _make_yandex_request(address)
            
        if not data:
            return None
        
        # Извлекаем координаты
        coords = _extract_best_result(data, address)
        latit=f'{coords[0]:.6f}'.replace(',','.')
        longit=f'{coords[1]:.6f}'.replace(',','.')
        if coords:
            #logger.info(f"Найдены координаты: {latit}, {longit}")
            return latit, longit
        else:
            #logger.warning("Координаты не найдены в московском регионе")
            return None
            
    except Exception as e:
        #logger.error(f"Критическая ошибка: {e}")
        return None

class App:
    def __init__(self, root):
        self.root = root
        root.title('Проставка широты и долготы')
        root.geometry('300x300')

        self.journal = tk.StringVar()

        tk.Label(root, text='Журнал:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.journal, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text="Выберите файл", command=self.select_journal).pack(padx=10, pady=5)

        tk.Button(root, text='Запустить обработку', command=self.run_processing, bg='#4CAF50', fg='white',font=('Segoe UI', 10, 'bold')).pack(padx=10, pady=15)

        self.log_text = scrolledtext.ScrolledText(root, wrap=tk.WORD, height=12, font=('Segoe UI', 9))
        self.log_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def select_journal(self):
        file = filedialog.askopenfilename(
            title='Выберите файл журнала',
            filetypes=[('Excel files', '*.xlsx')]
        )
        if file:
            self.journal.set(file)

    def run_processing(self):
        if not self.journal.get():
            messagebox.showerror('Ошибка', 'Выберите файл журнала!')
            return

        startingTime = datetime.now()
        self.log_text.delete(1.0, tk.END)
        self.log(f'=== Начало обработки: {startingTime.strftime("%H:%M:%S")} ===')

        try:
            # 1. Обработка
            self.process_files()
            endingTime = datetime.now()
            timeDif = endingTime - startingTime
            self.log(f'\n=== Готово! Всего: {timeDif.total_seconds():.1f} сек ===')
            messagebox.showinfo('Успех', 'Обработка завершена!\nРезультаты сохранены.')

        except Exception as e:
            self.log(f'\nКРИТИЧЕСКАЯ ОШИБКА: {e}')
            messagebox.showerror('Ошибка', f'Обработка прервана:\n{str(e)}')

    def process_files(self):
        journalPath = self.journal.get()
        # 2. Найти добавленные ранее и изменить
        try:
            # Загружаем журнал с помощью openpyxl, чтобы сохранить стили
            wb = ox.load_workbook(journalPath)
            ws = wb['Лист1']

            # Определяем индексы столбцов (1-based)
            headers = [cell.value for cell in ws[1]]
            col_indices = {}
            target_columns = [' Место ДТП (Адрес)', 'Координаты места ДТП (долгота)', 'Координаты места ДТП (широта)']
            for col_name in target_columns:
                if col_name in headers:
                    col_indices[col_name] = headers.index(col_name) + 1
                else:
                    self.log(f"Предупреждение: столбец '{col_name}' не найден в заголовке журнала.")
                    col_indices = {}
                    break
            API_KEY = '' #my_api
            if col_indices:
                # Обходим строки Excel начиная с 5751 (включительно)
                self.log(f'Будет проверено адресов:{ws.max_row + 1 - 5750}' )
                for row in range(5751, ws.max_row + 1):
                    # Извлекаем значения из Excel
                    addr_val= ws.cell(row=row, column=col_indices[' Место ДТП (Адрес)']).value
                    lat_val=ws.cell(row=row, column=col_indices['Координаты места ДТП (широта)']).value
                    # --- Координаты (оставляем пустыми) ---
                    lat, lon = '', ''
                    if not lat_val or str(lat_val).strip() == '':
                        #self.log(f'У адреса {addr_val} не прописаны координаты')
                        coords = get_coordinates_from_yandex(addr_val, API_KEY)
                        if coords:
                            lat, lon = coords
                            
                            ws.cell(row=row, column=col_indices['Координаты места ДТП (широта)'], value=lat)
                            ws.cell(row=row, column=col_indices['Координаты места ДТП (долгота)'], value=lon)
                        else:
                            self.log(f"Не удалось прочитать значение для адреса {addr_val}")
                        
                        nap(0.4)
                self.log(f"Прописаны координаты в безкоординатных строках (начиная с 5751).")

                # Сохраняем обновлённый файл
                output_path = journalPath[:-5] + '_newChanged.xlsx'
                wb.save(output_path)
                self.log(f"\nРезультат сохранён с сохранением форматирования: {output_path}")

        except Exception as e:
            self.log(f"Ошибка при обновлении журнала: {e}")

if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()