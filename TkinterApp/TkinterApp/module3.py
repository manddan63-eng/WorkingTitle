# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import pandas as pd
import numpy as np
from datetime import datetime, date, time as dtTime, timedelta
import re
import openpyxl as ox
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def rusLowerCaseWeekDays(num):
    days = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
    return days[num] if 0 <= num <= 6 else ''

def dateConvertion(someDate):
    """Возвращает datetime или None"""
    if isinstance(someDate, datetime):
        return someDate
    if isinstance(someDate, str):
        cleaned = re.sub(r'\s*[\.г\s]*$', '', someDate).strip()
        if cleaned:
            normalized = re.sub(r'[.\-]', '/', cleaned)
            try:
                return datetime.strptime(normalized, '%d/%m/%Y')
            except ValueError:
                return None
        else:
            return None
    return None

def timeCheck(value, fileName):
    if pd.isna(value) or value == '' or value is None:
        return ''
    if isinstance(value, str):
        val = value.strip()
        if not val:
            return ''
        if ':' not in val and '.' in val:
            parts = val.split('.')
            if len(parts) in (2, 3) and all(p.isdigit() for p in parts):
                val = ':'.join(parts)
        try:
            clean_val = re.sub(r'[^\d:]', '', val.split()[0]) if ' ' in val else val
            clean_val = clean_val.replace('AM', '').replace('PM', '').replace('am', '').replace('pm', '').strip()
            if ':' in clean_val:
                parts = clean_val.split(':')
                hour = int(parts[0])
                minute = int(parts[1]) if len(parts) > 1 else 0
                second = int(float(parts[2])) if len(parts) > 2 else 0
                if not (0 <= hour <= 23) or not (0 <= minute <= 59):
                    return ''
                return f"{hour:02d}:{minute:02d}:{second:02d}"
            elif val.isdigit():
                hour = int(val)
                if 0 <= hour <= 23:
                    return f"{hour:02d}:00:00"
            return ''
        except:
            return ''
    elif isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    elif isinstance(value, dtTime):
        dt = datetime.combine(datetime.min, value)
        return dt.strftime("%H:%M:%S")
    elif isinstance(value, (timedelta, pd.Timedelta)):
        total_seconds = int(round(value.total_seconds()))
        hours = total_seconds // 3600 % 24
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    else:
        return timeCheck(str(value), fileName)

def pddChapter(text):
    if pd.isna(text) or not isinstance(text, str):
        return ''
    match = re.search(r'\b(\d+(?:\s*[,\.]\s*\d+)+)', text)
    if match:
        normalized = re.sub(r'\s*,\s*', '.', match.group(1))
        normalized = re.sub(r'\s*\.\s*', '.', normalized)
        return normalized
    return ''

class App:
    def __init__(self, root):
        self.root = root
        root.title('Справки в список')
        root.geometry('500x550')

        self.inputFolder = tk.StringVar()
        self.journal = tk.StringVar()
        self.delayFile = tk.StringVar()

        tk.Label(root, text='Расположение папки со справками:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.inputFolder, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text='Выберите папку', command=self.select_inputFolder).pack(padx=10, pady=5)

        tk.Label(root, text='Журнал:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.journal, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text="Выберите файл", command=self.select_journal).pack(padx=10, pady=5)

        tk.Label(root, text='Расположение файла с задержками:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.delayFile, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text='Выберите файл задержек', command=self.select_delayFile).pack(padx=10, pady=5)

        tk.Button(root, text='Запустить обработку', command=self.run_processing, bg='#4CAF50', fg='white',font=('Segoe UI', 10, 'bold')).pack(padx=10, pady=15)

        self.log_text = scrolledtext.ScrolledText(root, wrap=tk.WORD, height=12, font=('Segoe UI', 9))
        self.log_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def select_inputFolder(self):
        folder = filedialog.askdirectory(title='Выберите папку со справками')
        if folder:
            self.inputFolder.set(folder)

    def select_journal(self):
        file = filedialog.askopenfilename(
            title='Выберите файл журнала',
            filetypes=[('Excel files', '*.xlsx')]
        )
        if file:
            self.journal.set(file)

    def select_delayFile(self):
        file = filedialog.askopenfilename(
            title='Выберите файл задержек',
            filetypes=[('Excel files', '*.xlsx'), ('Excel files (legacy)', '*.xls')]
        )
        if file:
            self.delayFile.set(file)

    def find_missing_reports(self, journal_df, delays_df, os_path):
        if delays_df.empty:
            self.log("Нет данных в файле задержек.")
            return pd.DataFrame(columns=['Гаражный номер'])

        if journal_df.empty:
            self.log("Журнал за 60 дней пуст — все машины без справок.")
            result = delays_df[['№ машины']].rename(columns={'№ машины': 'Гаражный номер'})
            # Преобразуем к int, если возможно
            result['Гаражный номер'] = pd.to_numeric(result['Гаражный номер'], errors='coerce')
            result = result.dropna()
            result['Гаражный номер'] = result['Гаражный номер'].astype(int)
            return result

        # --- Обработка журнала: только столкновения ---
        journal_collisions = journal_df[
            journal_df['Вид ДТП'].astype(str).str.contains('столкновен', case=False, na=False)
        ].copy()

        # Очистка и приведение к строке, затем к int
        journal_collisions['Гаражный номер'] = pd.to_numeric(
            journal_collisions['Гаражный номер'], errors='coerce'
        )
        journal_collisions = journal_collisions.dropna(subset=['Гаражный номер'])
        journal_collisions['Гаражный номер'] = journal_collisions['Гаражный номер'].astype(int)
        journal_collisions = journal_collisions[['Гаражный номер']].reset_index(drop=True)

        # Группировка по количеству справок
        journal_collisions['Кол-во справок за последние 60 дней'] = journal_collisions.groupby('Гаражный номер')['Гаражный номер'].transform('count')
        journal_collisions = journal_collisions.drop_duplicates(subset='Гаражный номер', keep='last')

        # --- Обработка файла задержек ---
        delays_df = delays_df.copy()
        delays_df['№ машины'] = pd.to_numeric(delays_df['№ машины'], errors='coerce')
        delays_df = delays_df.dropna(subset=['№ машины'])
        delays_df['№ машины'] = delays_df['№ машины'].astype(int)

        delays_df['Кол-во столкновений'] = delays_df.groupby('№ машины')['№ машины'].transform('count')
        delays_df = delays_df.drop_duplicates(subset='№ машины', keep='last')

        # --- Слияние ---
        result = pd.merge(
            delays_df,
            journal_collisions,
            left_on='№ машины',
            right_on='Гаражный номер',
            how='left'
        )

        # Заполняем отсутствующие значения нулями и приводим к int
        result['Кол-во справок за последние 60 дней'] = result['Кол-во справок за последние 60 дней'].fillna(0).astype(int)
        result['Кол-во столкновений'] = result['Кол-во столкновений'].astype(int)

        # Фильтрация: где столкновений больше, чем справок
        result = result[result['Кол-во столкновений'] > result['Кол-во справок за последние 60 дней']]

        # === УДАЛЯЕМ СТОЛБЕЦ "Причина" ===
        # Оставляем только нужные столбцы БЕЗ "Причина"
        result = result[['№ машины', 'Кол-во столкновений', 'Кол-во справок за последние 60 дней']].copy()
        result.rename(columns={'№ машины': 'Гаражный номер'}, inplace=True)

        self.log(f"Найдено {len(result)} машин без справок о столкновении.")
        return result
    def read_journal_last_60_days(self, journal_path ):
        """Возвращает (DataFrame, start_date, end_date)"""
        self.log(f"\nЧтение журнала: {os.path.basename(journal_path)}")
        try:
            journal_df = pd.read_excel(journal_path, sheet_name='Лист1')

            if 'Дата ДТП' not in journal_df.columns:
                self.log("В журнале нет столбца 'Дата ДТП'")
                today = pd.Timestamp.today().normalize()
                return pd.DataFrame(), today - pd.Timedelta(days=60), today

            journal_df['Дата ДТП'] = pd.to_datetime(journal_df['Дата ДТП'], format='%d.%m.%Y', errors='coerce')
            journal_df = journal_df.dropna(subset=['Дата ДТП'])

            today = pd.Timestamp.today().normalize()
            cutoff_date = today - pd.Timedelta(days=60)

            recent_df = journal_df[journal_df['Дата ДТП'] >= cutoff_date].copy()
            self.log(f"Найдено {len(recent_df)} записей за период: {cutoff_date.strftime('%d.%m.%Y')} – {today.strftime('%d.%m.%Y')}")
            return recent_df, cutoff_date, today

        except Exception as e:
            self.log(f"Ошибка при чтении журнала: {e}")
            today = pd.Timestamp.today().normalize()
            return pd.DataFrame(), today - pd.Timedelta(days=60), today

    def run_processing(self):
        """if not self.inputFolder.get():
            messagebox.showerror('Ошибка', 'Выберите папку со справками!')
            return
        if not self.journal.get():
            messagebox.showerror('Ошибка', 'Выберите файл журнала!')
            return
        if not self.delayFile.get():
            messagebox.showerror('Ошибка', 'Выберите файл задержек!')
            return"""

        startingTime = datetime.now()
        self.log_text.delete(1.0, tk.END)
        self.log(f'=== Начало обработки: {startingTime.strftime("%H:%M:%S")} ===')

        try:
            # 1. Обработка справок → обновлённый журнал
            #self.process_files()

            # 2. Получаем данные из задержек
            #delays_df = self.process_delay_file()

            # 3. Читаем обновлённый журнал за 60 дней
            #updated_journal_path = self.journal.get()[:-5] + '_new.xlsx'
            #journal_recent, start_date, end_date = self.read_journal_last_60_days(updated_journal_path)
            


            # 4. Находим машины без справок
            #missing_df = self.find_missing_reports(journal_recent, delays_df,updated_journal_path )

            # 5. Сохраняем результат
            """if not missing_df.empty:
                # Формат дат: dd-mm-yy
                start_str = start_date.strftime('%d-%m-%y')
                end_str = end_date.strftime('%d-%m-%y')
                output_filename = f"Гаражные_номера_без_справок_{start_str}_{end_str}.xlsx"
                output_path = os.path.join(os.path.dirname(updated_journal_path), output_filename)
                missing_df.to_excel(output_path, index=False)
                self.log(f"\nСохранён список без справок: {output_filename}")
            else:
                self.log("\nВсе машины из задержек имеют справки о столкновении.")
"""
            #updated_journal_path = self.journal.get()
            updated_journal_path='C:/Users/Danara/Downloads/Журнал с дтп 2025 для дашборта_new.xlsx'
            self.generate_weekly_dashboard_2025(updated_journal_path)
            endingTime = datetime.now()
            timeDif = endingTime - startingTime
            self.log(f'\n=== Готово! Всего: {timeDif.total_seconds():.1f} сек ===')
            messagebox.showinfo('Успех', 'Обработка завершена!\nРезультаты сохранены.')

        except Exception as e:
            self.log(f'\nКРИТИЧЕСКАЯ ОШИБКА: {e}')
            messagebox.showerror('Ошибка', f'Обработка прервана:\n{str(e)}')

    def process_delay_file(self):
        delay_path = self.delayFile.get()
        self.log(f"\nОбработка файла задержек: {os.path.basename(delay_path)}")

        try:
            # Выбор движка в зависимости от расширения
            if delay_path.lower().endswith('.xls'):
                engine = 'xlrd'
            else:
                engine = 'openpyxl'

            file_df = pd.read_excel(delay_path, header=None, engine=engine)

            file2 = file_df.iloc[:, [11, 13, 19]].copy()
            file2.columns = ['col11', 'Unnamed: 13', 'Unnamed: 19']

            file2['№ машины'] = file2['Unnamed: 13'].shift(-4)
            file2.rename(columns={'Unnamed: 19': 'Причина'}, inplace=True)

            mask = file2['Причина'].notna() & file2['Причина'].astype(str).str.contains('Столкновение', na=False)
            file2 = file2[mask].copy()
            file2['Причина'] = file2['Причина'].astype(str).str.replace('Причина:', '', regex=False).str.strip()

            result_df = file2[['Причина', '№ машины']].reset_index(drop=True)
            self.log(f"Загружено {len(result_df)} записей из файла задержек.")
            #записать!!

            output_filename = f"Гаражные_номера_без_справок3.xlsx"
            output_path = os.path.join('C:/Users/Danara/Downloads/', output_filename)
            result_df.to_excel(output_path, index=False)

            return result_df

        except Exception as e:
            self.log(f"Ошибка при обработке файла задержек: {e}")
            return pd.DataFrame()

    def process_files(self):
        folderPath = self.inputFolder.get()
        journalPath = self.journal.get()

        items = [f for f in os.listdir(folderPath) if not f.startswith(('Журнал', 'Задержк'))]
        total_files = len(items)

        if total_files == 0:
            self.log("Нет файлов для обработки (файлы 'Журнал','Задержки' игнорируются).")
            return

        Result = pd.DataFrame()
        failed_files = []
        empty_fields_warnings = []

        for idx, item in enumerate(items, start=1):
            filePath = os.path.join(folderPath, item)
            self.log(f"[{idx}/{total_files}] Обработка: {item}")

            try:
                if not filePath.endswith('.xls'):
                    try:
                        dataFile = pd.read_excel(filePath, engine='openpyxl').fillna('')
                    except Exception:
                        dataFile = pd.read_excel(filePath, engine='xlrd').fillna('')
                else:
                    dataFile = pd.read_excel(filePath, engine='xlrd').fillna('')

                # --- Дата и день недели ---
                original_date_cell = dataFile.iloc[1][2]
                converted_date = None
                weekDay = ''
                if original_date_cell != '' and not pd.isna(original_date_cell):
                    try:
                        converted_date = dateConvertion(original_date_cell)
                        if converted_date:
                            weekDay = rusLowerCaseWeekDays(converted_date.weekday())
                    except Exception:
                        weekDay = ''

                # --- ФИО водителя ---
                p = str(dataFile.iloc[18][6]).split(' ')
                shortName = f"{p[0]} {p[1][0].upper()+'.' if len(p) > 1 and p[1] else ''}{p[2][0].upper()+'.' if len(p) > 2 and p[2] else ''}".strip()

                # --- Пункт правил ---
                original_pdd_text = dataFile.iloc[40][1]
                pdd_value = ''
                if isinstance(original_pdd_text, str):
                    textLower = original_pdd_text.lower().strip()
                    hasPDD = 'пдд' in textLower
                    hasP = re.search(r'\bп\.?\s*\d', textLower) is not None
                    if hasPDD or hasP:
                        pdd_value = pddChapter(original_pdd_text)

                # --- Стаж ---
                raw_stazh = str(dataFile.iloc[24][9]).strip()
                raw_stazh_park = str(dataFile.iloc[24][20]).strip()
                def process_stazh(raw):
                    if not raw or raw.lower() in ('', 'nan'):
                        return ''
                    if re.fullmatch(r'[\d\s.,]+', raw) and re.search(r'\d', raw):
                        num_part = re.search(r'\d+', raw)
                        if num_part:
                            return int(num_part.group())
                    return raw

                stazh_obshch = process_stazh(raw_stazh)
                stazh_v_parke = process_stazh(raw_stazh_park)

                # --- Пострадавшие ---
                postradavshie_raw = dataFile.iloc[14][8]
                postradavshie = 0
                if postradavshie_raw != '' and str(postradavshie_raw).lower() not in ('нет', ''):
                    try:
                        postradavshie = int(postradavshie_raw)
                    except:
                        postradavshie = 0

                # --- Координаты (оставляем пустыми) ---
                lat, lon = '', ''

                # --- DataFrame ---
                dfData = pd.DataFrame({
                    'Место': '1' if 'европротокол' in str(dataFile.iloc[66][7]).lower() else '',
                    'Дата ДТП': converted_date,
                    'Время ДТП': timeCheck(dataFile.iloc[1][17], item),
                    'День недели': weekDay,
                    ' Место ДТП (Адрес)': dataFile.iloc[9][6],
                    'Район': dataFile.iloc[10][9],
                    'Округ': dataFile.iloc[10][2],
                    'Координаты места ДТП (широта)': lat,
                    'Координаты места ДТП (долгота)': lon,
                    '3-я сторона': '',
                    'Название стороней организации': dataFile.iloc[74][6],
                    'Государственный регистрационный знак сторонего транспорта': dataFile.iloc[73][6],
                    'МГТ': '',
                    'Перевозчик': 'ГУП "Мосгортранс"',
                    'Название филиала': str(dataFile.iloc[17][4])
                        .replace('Филиал ', 'Ф').replace('Филилал ', 'Ф')
                        .replace('Южный', 'Ю').replace('Северный', 'С')
                        .replace('Юго-', 'Ю').replace('Северо-', 'С')
                        .replace('Восточный', 'В').replace('Западный', 'З'),
                    'Название площадки': dataFile.iloc[17][15],
                    'Маршрут': dataFile.iloc[31][8],
                    'Марка автобуса / электробуса': dataFile.iloc[32][8],
                    'Гаражный номер': '' if dataFile.iloc[33][20] == '' else int(dataFile.iloc[33][20]),
                    'Регистрационный номер': dataFile.iloc[33][8],
                    'Водитель': '',
                    'Табельный номер водителя': dataFile.iloc[19][6],
                    'ФИО водителя': shortName,
                    'Гражданство': dataFile.iloc[22][6],
                    'Возраст': dataFile.iloc[21][6],
                    'Стаж общий ': stazh_obshch,
                    'Стаж в парке': stazh_v_parke,
                    'ДТП': '',
                    'Вид ДТП': dataFile.iloc[7][3],
                    'Причина ДТП': dataFile.iloc[40][1],
                    'Виновник ДТП': str(dataFile.iloc[65][9])
                        .replace('Не вина', '3-е лицо')
                        .replace('Вина', 'Перевозчик')
                        .replace('В расследовании', 'Проводится разбор'),
                    'Пункт правил': pdd_value,
                    'Скорость по гланассу КМ/Ч': dataFile.iloc[36][16],
                    'Выделенная полоса (ДА; НЕТ,)': str(dataFile.iloc[12][9]).lower(),
                    'Пострадавшие': '',
                    'Кол-во пострадавших': postradavshie,
                    'в т.ч.    лёгкий   вред здоровью': postradavshie,
                    'в т.ч. средний вред здоровью': '0',
                    'в т.ч. тяжёлый вред здоровью': '0',
                    'Кол-во погибших': '0',
                    'ГК': '0',
                    'Ответст-сть': '',
                    'Постановление': dataFile.iloc[66][7],
                    'Дата постановления': (dateConvertion(dataFile.iloc[67][10])).strftime('%d.%m.%Y') if dateConvertion(dataFile.iloc[67][10]) else '',
                    'Наказание водителя': 'выговор' if dataFile.iloc[65][9] == 'Вина' else '',
                    'Проишествия': '',
                    'Резонансные проишествия': '',
                    'Проишествия с водителями': '',
                    'Проишествия с контролёрами': '',
                    'Проишествия с пассажирами': '',
                    'Кол-во задержек в движении': '',
                    'Сработка АНТИСОН': '',
                    'Кол-во отстранённых водителей.': '',
                    'Проишествия3': ''
                }, index=[0])

                # --- Проверка потерь данных ---
                fields_to_check = {
                    'Дата ДТП': original_date_cell,
                    'Время ДТП': dataFile.iloc[1][17],
                    'ФИО водителя': dataFile.iloc[18][6],
                    'Пункт правил': original_pdd_text,
                    'Стаж общий ': raw_stazh,
                    'Стаж в парке': raw_stazh_park,
                    'Кол-во пострадавших': postradavshie_raw,
                }

                for col_name, orig_val in fields_to_check.items():
                    final_val = dfData[col_name].iloc[0]
                    if (
                        not pd.isna(orig_val) and str(orig_val).strip() != '' and
                        (pd.isna(final_val) or final_val == '')
                    ):
                        empty_fields_warnings.append(f"{item} -> {col_name}: '{orig_val}' -> ''")

                Result = pd.concat([Result, dfData], ignore_index=True)

            except Exception as e:
                failed_files.append(item)
                self.log(f"[{idx}/{total_files}] Ошибка: {e}")

        # --- УДАЛЕНИЕ ДУБЛИКАТОВ ---
        if Result.empty:
            self.log("Нет данных для сохранения.")
            return

        key_cols = ['Гаражный номер', 'Дата ДТП', 'Время ДТП', 'Маршрут']

        for col in key_cols:
            if col not in Result.columns:
                Result[col] = ''

        # 1. Удалить дубликаты внутри новых записей
        Result_before = len(Result)
        Result = Result.drop_duplicates(subset=key_cols, keep='first')
        removed_internal = Result_before - len(Result)
        if removed_internal > 0:
            self.log(f"Удалено {removed_internal} дубликатов внутри новых записей.")

        # 2. Удалить записи, уже существующие в исходном журнале
        try:
            existing_journal = pd.read_excel(journalPath, sheet_name='Лист1')
            # Привести дату к строке в формате dd.mm.yyyy
            if 'Дата ДТП' in existing_journal.columns:
                existing_journal['Дата ДТП'] = pd.to_datetime(
                    existing_journal['Дата ДТП'], format='%d.%m.%Y', errors='coerce'
                ).dt.strftime('%d.%m.%Y').fillna('')
            else:
                existing_journal['Дата ДТП'] = ''
            # Убедиться, что все ключевые столбцы есть
            for col in key_cols:
                if col not in existing_journal.columns:
                    existing_journal[col] = ''
        except Exception as e:
            self.log(f"Не удалось прочитать существующий журнал: {e}")
            existing_journal = pd.DataFrame()

        if not existing_journal.empty:
            # Создаём хэши строк для сравнения
            existing_keys = set(
                existing_journal[key_cols].astype(str).apply('|'.join, axis=1)
            )
            result_keys = Result[key_cols].astype(str).apply('|'.join, axis=1)

            # Фильтруем
            Result_before = len(Result)
            Result = Result[~result_keys.isin(existing_keys)]
            removed_existing = Result_before - len(Result)
            if removed_existing > 0:
                self.log(f"Пропущено {removed_existing} записей — уже есть в журнале.")
        else:
            self.log(f"Новых записей без дубликатов: {len(Result)}")

        # --- Сортировка по дате ---
        if not Result.empty and 'Дата ДТП' in Result.columns:
            Result = Result.sort_values(by='Дата ДТП', na_position='last').reset_index(drop=True)

        # --- Сохранение в Excel ---
        output_path = journalPath[:-5] + '_new.xlsx'
        workbook = ox.load_workbook(journalPath)
        sheet = workbook['Лист1']

        filledColNums = [2, 3, 4, 5, 6, 7, 11, 12, 15, 16, 17, 18, 19, 20, 22, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 43, 44]
        fillingBorder = Side(border_style='thin', color='AFC69F')
        fillingEmpty = PatternFill(fill_type='solid', start_color='ffffc000', end_color='ffffc000')

        for i, r in enumerate(dataframe_to_rows(Result, index=False, header=False)):
            clean_row = []
            for val in r:
                if val is None or (isinstance(val, float) and np.isnan(val)) or str(val).strip() == '':
                    clean_row.append('')
                else:
                    clean_row.append(val)

            sheet.append(clean_row)
            lastRowNew = sheet.max_row

            fillingStyle = PatternFill(
                fill_type='solid',
                start_color='ffc5e0b3' if i % 2 != 0 else 'ffe2efd9',
                end_color='ffc5e0b3' if i % 2 != 0 else 'ffe2efd9'
            )

            for cell in sheet[lastRowNew]:
                col_idx = cell.column
                val = cell.value
                if col_idx in (2, 43):
                    if isinstance(val, (datetime, date)) and not isinstance(val, str):
                        cell.number_format = 'DD.MM.YYYY'

                if cell.value == '' and col_idx in filledColNums:
                    cell.fill = fillingEmpty
                else:
                    cell.fill = fillingStyle

                cell.border = Border(left=fillingBorder, right=fillingBorder, top=fillingBorder, bottom=fillingBorder)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(output_path)
        self.log(f"\nРезультат сохранён: {output_path}")

        # --- Итоговый отчёт ---
        if failed_files:
            self.log("\nНеобработанные файлы:")
            for f in failed_files:
                self.log(f"  - {f}")
        else:
            self.log("\nВсе файлы обработаны.")

        if empty_fields_warnings:
            self.log(f"\nПотери данных ({len(empty_fields_warnings)} записей):")
            for w in empty_fields_warnings[:50]:
                self.log(f"  - {w}")
            if len(empty_fields_warnings) > 50:
                self.log(f"  ... и ещё {len(empty_fields_warnings) - 50}")
        else:
            self.log("\nНет потерь данных.")

    def generate_weekly_dashboard_2025(self, journal_path):
        try:


            self.log("Генерация дашборда за 2025 год...")
            df = pd.read_excel(journal_path, sheet_name='Лист1')

            df['Дата ДТП'] = pd.to_datetime(df['Дата ДТП'], format='%d.%m.%Y', errors='coerce')
            df = df.dropna(subset=['Дата ДТП'])

            df_2025 = df[df['Дата ДТП'].dt.year == 2025].copy()
            df_2025['Неделя'] = df_2025['Дата ДТП'].dt.isocalendar().week

            total_all = df_2025.groupby('Неделя').size()
            carrier_all = df_2025[df_2025['Виновник ДТП'] == 'Перевозчик'].groupby('Неделя').size()
            df1 = pd.concat([total_all, carrier_all], axis=1).fillna(0).astype(int)
            df1.columns = ['Всего ДТП', 'По вине перевозчика']
            df1 = df1.reset_index()

            df_with_victims = df_2025[df_2025['Кол-во пострадавших'] != 0]
            total_victims = df_with_victims.groupby('Неделя').size()
            carrier_victims = df_with_victims[df_with_victims['Виновник ДТП'] == 'Перевозчик'].groupby('Неделя').size()
            df2 = pd.concat([total_victims, carrier_victims], axis=1).fillna(0).astype(int)
            df2.columns = ['Всего ДТП', 'По вине перевозчика']
            df2 = df2.reset_index()

            df_without_victims = df_2025[df_2025['Кол-во пострадавших'] == 0]
            total_no_victims = df_without_victims.groupby('Неделя').size()
            carrier_no_victims = df_without_victims[df_without_victims['Виновник ДТП'] == 'Перевозчик'].groupby('Неделя').size()
            df3 = pd.concat([total_no_victims, carrier_no_victims], axis=1).fillna(0).astype(int)
            df3.columns = ['Всего ДТП', 'По вине перевозчика']
            df3 = df3.reset_index()

            weeks1 = df1['Неделя'].tolist()
            total1 = df1['Всего ДТП'] .tolist()
            rest1 = (df1['Всего ДТП'] - df1['По вине перевозчика']).tolist()
            carrier1 = df1['По вине перевозчика'].tolist()

            weeks2 = df2['Неделя'].tolist()
            total2 = df2['Всего ДТП'] .tolist()
            rest2 = (df2['Всего ДТП'] - df2['По вине перевозчика']).tolist()
            carrier2 = df2['По вине перевозчика'].tolist()

            weeks3 = df3['Неделя'].tolist()
            total3 = df3['Всего ДТП'] .tolist()
            rest3 = (df3['Всего ДТП'] - df3['По вине перевозчика']).tolist()
            carrier3 = df3['По вине перевозчика'].tolist()



            self.log("Генерация дашборда за 2026 год...")

            df_2026 = df[df['Дата ДТП'].dt.year == 2026].copy()
            df_2026['Неделя'] = df_2026['Дата ДТП'].dt.isocalendar().week

            total_all_2026 = df_2026.groupby('Неделя').size()
            carrier_all_2026 = df_2026[df_2026['Виновник ДТП'] == 'Перевозчик'].groupby('Неделя').size()
            df1_2026 = pd.concat([total_all_2026, carrier_all_2026], axis=1).fillna(0).astype(int)
            df1_2026.columns = ['Всего ДТП', 'По вине перевозчика']
            df1_2026 = df1_2026.reset_index()

            df_with_victims_2026 = df_2026[df_2026['Кол-во пострадавших'] != 0]
            total_victims_2026 = df_with_victims_2026.groupby('Неделя').size()
            carrier_victims_2026 = df_with_victims_2026[df_with_victims_2026['Виновник ДТП'] == 'Перевозчик'].groupby('Неделя').size()
            df2_2026 = pd.concat([total_victims_2026, carrier_victims_2026], axis=1).fillna(0).astype(int)
            df2_2026.columns = ['Всего ДТП', 'По вине перевозчика']
            df2_2026 = df2_2026.reset_index()

            df_without_victims_2026 = df_2026[df_2026['Кол-во пострадавших'] == 0]
            total_no_victims_2026 = df_without_victims_2026.groupby('Неделя').size()
            carrier_no_victims_2026 = df_without_victims_2026[df_without_victims_2026['Виновник ДТП'] == 'Перевозчик'].groupby('Неделя').size()
            df3_2026 = pd.concat([total_no_victims_2026, carrier_no_victims_2026], axis=1).fillna(0).astype(int)
            df3_2026.columns = ['Всего ДТП', 'По вине перевозчика']
            df3_2026 = df3_2026.reset_index()

            weeks1_2026 = df1_2026['Неделя'].tolist()
            total1_2026 = df1_2026['Всего ДТП'] .tolist()
            rest1_2026 = (df1_2026['Всего ДТП'] - df1_2026['По вине перевозчика']).tolist()
            carrier1_2026 = df1_2026['По вине перевозчика'].tolist()

            weeks2_2026 = df2_2026['Неделя'].tolist()
            total2_2026 = df2_2026['Всего ДТП'] .tolist()
            rest2_2026 = (df2_2026['Всего ДТП'] - df2_2026['По вине перевозчика']).tolist()
            carrier2_2026 = df2_2026['По вине перевозчика'].tolist()

            weeks3_2026 = df3_2026['Неделя'].tolist()
            total3_2026 = df3_2026['Всего ДТП'] .tolist()
            rest3_2026 = (df3_2026['Всего ДТП'] - df3_2026['По вине перевозчика']).tolist()
            carrier3_2026 = df3_2026['По вине перевозчика'].tolist()

            html_content = f'''<!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <title>ДТП по неделям — 2025/2026</title>
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; background: #f9f9f9; }}
            .container {{ max-width: 1000px; margin: 0 auto; }}
            h1 {{ text-align: center; color: #2c3e50; }}
            .overallStyle {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); margin-bottom: 30px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>ДТП по неделям — 2025/2026 год</h1>
            <div id="chart" class="overallStyle"></div>
        </div>

        <div class="container">
            <h1>ДТП по вине перевозчика — с пострадавшими</h1>
            <div id="chart2" class="overallStyle"></div>
        </div>

        <div class="container">
            <h1>ДТП по вине перевозчика — без пострадавших</h1>
            <div id="chart3" class="overallStyle"></div>
        </div>
    <script>
        const weeks1 = {weeks1};
        const total1 = {total1};
        const carrier1 = {carrier1};
        const rest1 = {rest1};

        const weeks2 = {weeks2};
        const total2 = {total2};
        const carrier2 = {carrier2};
        const rest2 = {rest2};

        const weeks3 = {weeks3};
        const total3 = {total3};
        const carrier3 = {carrier3};
        const rest3 = {rest3};


        const weeks1_2026 = {weeks1_2026};
        const total1_2026 = {total1_2026};
        const carrier1_2026 = {carrier1_2026};
        const rest1_2026 = {rest1_2026};

        const weeks2_2026 = {weeks2_2026};
        const total2_2026 = {total2_2026};
        const carrier2_2026 = {carrier2_2026};
        const rest2_2026 = {rest2_2026};

        const weeks3_2026 = {weeks3_2026};
        const total3_2026 = {total3_2026};
        const carrier3_2026 = {carrier3_2026};
        const rest3_2026 = {rest3_2026};

        // сдвинутые координаты для группировки
        const x_2025_1 = weeks1.map(w => w - 0.15);
        const x_2026_1 = weeks1_2026.map(w => w + 0.15);

        Plotly.newPlot('chart', [
            // === 2025 год ===
            {{
                x: x_2025_1,
                y: carrier1,
                type: 'bar',
                name: '2025: По вине перевозчика',
                marker: {{ color: '#C73E1D' }},
                text: carrier1,
                textposition: 'inside',
                textangle: 0,
                insidetextanchor: 'middle',
                textfont: {{ color: 'white', size: 12 }},
                customdata:  total1,
                hovertemplate: 'По вине перевозчика: %{{y}}<extra></extra>'
            }},
            {{
                x: x_2025_1,
                y: rest1,
                type: 'bar',
                name: '2025: Всего ДТП',
                marker: {{ color: '#2E86AB' }},
                text: total1.map(val => `2025: ${{val}}`),
                textposition: 'outside',
                textangle: 0,
                textfont: {{ color: '#2E86AB', size: 12 }},
                customdata:  total1,
                hovertemplate: 'Все ДТП: %{{customdata}}<extra></extra>'
            }},

            // === 2026 год ===
            {{
                x: x_2026_1,
                y: carrier1_2026,
                type: 'bar',
                name: '2026: По вине перевозчика',
                marker: {{ color: '#A52A2A' }},
                text: carrier1_2026,
                textposition: 'inside',
                textangle: 0,
                insidetextanchor: 'middle',
                textfont: {{ color: 'white', size: 12 }},
                customdata:  total1_2026,
                hovertemplate: 'По вине перевозчика: %{{y}}<extra></extra>'
            }},
            {{
                x: x_2026_1,
                y: rest1_2026,
                type: 'bar',
                name: '2026: Всего ДТП',
                marker: {{ color: '#1E5E8C' }},
                text: total1_2026.map(val => `2026: ${{val}}`),
                textposition: 'outside',
                textangle: 0,
                textfont: {{ color: '#1E5E8C', size: 12 }},
                customdata:  total1_2026,
                hovertemplate: 'Все ДТП: %{{customdata}}<extra></extra>'
            }}
        ], {{
            title: 'ДТП по неделям',
            xaxis: {{
                title: 'Неделя года',
                tickmode: 'array',
                tickvals: Array.from({{length: 53}}, (_, i) => i + 1),
                range: [0.0, 5.0]
            }},
            yaxis: {{ title: 'Количество ДТП' }},
            barmode: 'stack',
            hovermode: 'x unified',
            legend: {{ orientation: 'h', yanchor: 'bottom', y: 1.02, xanchor: 'right', x: 1 }}
        }});

        // сдвинутые координаты для группировки
        const x_2025_2 = weeks2.map(w => w - 0.15);
        const x_2026_2 = weeks2_2026.map(w => w + 0.15);

        Plotly.newPlot('chart2', [
            // === 2025 год ===
            {{
                x: x_2025_2,
                y: carrier2,
                type: 'bar',
                name: '2025: По вине перевозчика',
                marker: {{ color: '#C73E1D' }},
                text: carrier2,
                textposition: 'inside',
                textangle: 0,
                insidetextanchor: 'middle',
                textfont: {{ color: 'white', size: 12 }},
                customdata: total2,
                hovertemplate: 'По вине перевозчика: %{{y}}<extra></extra>'
            }},
            {{
                x: x_2025_2,
                y: rest2,
                type: 'bar',
                name: '2025: Всего ДТП',
                marker: {{ color: '#2E86AB' }},
                text: total2.map(val => `2025: ${{val}}`),
                textposition: 'outside',
                textangle: 0,
                textfont: {{ color: '#2E86AB', size: 12 }},
                customdata: total2,
                hovertemplate: 'Все ДТП: %{{customdata}}<extra></extra>'
            }},

            // === 2026 год ===
            {{
                x: x_2026_2,
                y: carrier2_2026,
                type: 'bar',
                name: '2026: По вине перевозчика',
                marker: {{ color: '#A52A2A' }},
                text: carrier2_2026,
                textposition: 'inside',
                textangle: 0,
                insidetextanchor: 'middle',
                textfont: {{ color: 'white', size: 12 }},
                customdata: total2_2026,
                hovertemplate: 'По вине перевозчика: %{{y}}<extra></extra>'
            }},
            {{
                x: x_2026_2,
                y: rest2_2026,
                type: 'bar',
                name: '2026: Всего ДТП',
                marker: {{ color: '#1E5E8C' }},
                text: total2_2026.map(val => `2026: ${{val}}`),
                textposition: 'outside',
                textangle: 0,
                textfont: {{ color: '#1E5E8C', size: 12 }},
                customdata: total2_2026,
                hovertemplate: 'Все ДТП: %{{customdata}}<extra></extra>'
            }}
        ], {{
            title: 'ДТП по вине перевозчика — с пострадавшими',
            xaxis: {{
                title: 'Неделя года',
                tickmode: 'array',
                tickvals: Array.from({{length: 53}}, (_, i) => i + 1),
                range: [0.0, 5.0]
            }},
            yaxis: {{ title: 'Количество ДТП' }},
            barmode: 'stack',
            hovermode: 'x unified'
        }});


        const x_2025_3 = weeks3.map(w => w - 0.15);
        const x_2026_3 = weeks3_2026.map(w => w + 0.15);
        Plotly.newPlot('chart3', [
            // === 2025 год ===
            {{
                x: x_2025_3,
                y: carrier3,
                type: 'bar',
                name: '2025: По вине перевозчика',
                marker: {{ color: '#C73E1D' }},
                text: carrier3,
                textposition: 'inside',
                textangle: 0,
                insidetextanchor: 'middle',
                textfont: {{ color: 'white', size: 12 }},
                customdata: total3,
                hovertemplate: 'По вине перевозчика: %{{y}}<extra></extra>'
            }},
            {{
                x: x_2025_3,
                y: rest3,
                type: 'bar',
                name: '2025: Всего ДТП',
                marker: {{ color: '#2E86AB' }},
                text: total3.map(val => `2025: ${{val}}`),
                textposition: 'outside',
                textangle: 0,
                textfont: {{ color: '#2E86AB', size: 12 }},
                customdata: total3,
                hovertemplate: 'Все ДТП: %{{customdata}}<extra></extra>'
            }},

            // === 2026 год ===
            {{
                x: x_2026_3,
                y: carrier3_2026,
                type: 'bar',
                name: '2026: По вине перевозчика',
                marker: {{ color: '#A52A2A' }},
                text: carrier3_2026,
                textposition: 'inside',
                textangle: 0,
                insidetextanchor: 'middle',
                textfont: {{ color: 'white', size: 12 }},
                customdata: total3_2026,
                hovertemplate: 'По вине перевозчика: %{{y}}<extra></extra>'
            }},
            {{
                x: x_2026_3,
                y: rest3_2026,
                type: 'bar',
                name: '2026: Всего ДТП',
                marker: {{ color: '#1E5E8C' }},
                text: total3_2026.map(val => `2026: ${{val}}`),
                textposition: 'outside',
                textangle: 0,
                textfont: {{ color: '#1E5E8C', size: 12 }},
                customdata: total3_2026,
                hovertemplate: 'Все ДТП: %{{customdata}}<extra></extra>'
            }}
        ], {{
            title: 'ДТП по вине перевозчика — без пострадавших',
            xaxis: {{
                title: 'Неделя года',
                tickmode: 'array',
                tickvals: Array.from({{length: 53}}, (_, i) => i + 1),
                range: [0.0, 5.0]
            }},
            yaxis: {{ title: 'Количество ДТП' }},
            barmode: 'stack',
            hovermode: 'x unified'
        }});
    </script>
    </body>
    </html>'''

            output_html = journal_path.replace('.xlsx', '_dashboard_2025_weekly.html')
            with open(output_html, 'w', encoding='utf-8') as f:
                f.write(html_content)

            self.log(f"Дашборд за 2025 год сохранён: {os.path.basename(output_html)}")

        except Exception as e:
            self.log(f"Ошибка при создании дашборда: {e}")


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()
