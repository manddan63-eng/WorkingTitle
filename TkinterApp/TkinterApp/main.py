# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import pandas as pd
import numpy as np
from datetime import datetime, date, time as dtTime, timedelta
import re
import openpyxl as ox
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import time

# === ЗАМЕНИТЕ ЭТОТ КЛЮЧ НА СВОЙ ИЗ YANDEX.CLOUD ===
YANDEX_API_KEY = '8ae5ed96-1e34-48a7-acbf-878c5553db25'


def yandex_geocode2(address, api_key):
    url = "https://geocode-maps.yandex.ru/1.x/"
    params = {
        "apikey": api_key,          # ← обязательно!
        "geocode": address,
        "format": "json",
        "lang": "ru_RU"
    }
    try:
        response = requests.get(url, params=params, timeout=10)
        if response.status_code == 200:
            data = response.json()
            try:
                found = int(data['response']['GeoObjectCollection']['metaDataProperty']['GeocoderResponseMetaData']['found'])
                if found > 0:
                    pos = data['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']
                    lon, lat = pos.split()
                    return str(round(float(lat), 6)), str(round(float(lon), 6))
                else:
                    return '', ''
            except (KeyError, IndexError, ValueError):
                return '', ''
        else:
            print(f"Yandex API error: {response.status_code}")
            return '', ''
    except Exception as e:
        print(f"Exception: {e}")
        return '', ''

def yandex_geocode(address):
    url = "https://geocode-maps.yandex.ru/1.x/"
    params = {
        "geocode": address,
        "format": "json",
        "lang": "ru_RU"
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    }
    try:
        response = requests.get(url, params=params, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            try:
                found = int(data['response']['GeoObjectCollection']['metaDataProperty']['GeocoderResponseMetaData']['found'])
                if found > 0:
                    pos = data['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']
                    lon, lat = pos.split()
                    return str(round(float(lat), 6)), str(round(float(lon), 6))
                else:
                    return '', ''
            except (KeyError, IndexError, ValueError):
                return '', ''
        else:
            return '', ''
    except Exception:
        return '', ''
def rusLowerCaseWeekDays(num):
    days = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
    return days[num] if 0 <= num <= 6 else ''

def dateConvertion(someDate):
    """Возвращает datetime или None"""
    if isinstance(someDate, datetime):
        return someDate
    if isinstance(someDate, str):
        cleaned = re.sub(r'\s*[\.г\s]*$', '', someDate).strip()
        if cleaned:
            normalized = re.sub(r'[.\-]', '/', cleaned)
            try:
                return datetime.strptime(normalized, '%d/%m/%Y')
            except ValueError:
                return None
        else:
            return None
    return None

def timeCheck(value, fileName):
    if pd.isna(value) or value == '' or value is None:
        return ''
    if isinstance(value, str):
        val = value.strip()
        if not val:
            return ''
        if ':' not in val and '.' in val:
            parts = val.split('.')
            if len(parts) in (2, 3) and all(p.isdigit() for p in parts):
                val = ':'.join(parts)
        try:
            clean_val = re.sub(r'[^\d:]', '', val.split()[0]) if ' ' in val else val
            clean_val = clean_val.replace('AM', '').replace('PM', '').replace('am', '').replace('pm', '').strip()
            if ':' in clean_val:
                parts = clean_val.split(':')
                hour = int(parts[0])
                minute = int(parts[1]) if len(parts) > 1 else 0
                second = int(float(parts[2])) if len(parts) > 2 else 0
                if not (0 <= hour <= 23) or not (0 <= minute <= 59):
                    return ''
                return f"{hour:02d}:{minute:02d}:{second:02d}"
            elif val.isdigit():
                hour = int(val)
                if 0 <= hour <= 23:
                    return f"{hour:02d}:00:00"
            return ''
        except:
            return ''
    elif isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    elif isinstance(value, dtTime):
        dt = datetime.combine(datetime.min, value)
        return dt.strftime("%H:%M:%S")
    elif isinstance(value, (timedelta, pd.Timedelta)):
        total_seconds = int(round(value.total_seconds()))
        hours = total_seconds // 3600 % 24
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    else:
        return timeCheck(str(value), fileName)

def pddChapter(text):
    if pd.isna(text) or not isinstance(text, str):
        return ''
    match = re.search(r'\b(\d+(?:\s*[,\.]\s*\d+)+)', text)
    if match:
        normalized = re.sub(r'\s*,\s*', '.', match.group(1))
        normalized = re.sub(r'\s*\.\s*', '.', normalized)
        return normalized
    return ''

class App:
    def __init__(self, root):
        self.root = root
        root.title('Справки в список')
        root.geometry('500x550')

        self.inputFolder = tk.StringVar()
        self.journal = tk.StringVar()

        tk.Label(root, text='Расположение папки со справками:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.inputFolder, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text='Выберите папку', command=self.select_inputFolder).pack(padx=10, pady=5)

        tk.Label(root, text='Журнал:', font=('Segoe UI', 9)).pack(anchor='w', padx=10, pady=(10, 0))
        tk.Entry(root, textvariable=self.journal, width=50, state='readonly').pack(padx=10, pady=5)
        tk.Button(root, text="Выберите файл", command=self.select_journal).pack(padx=10, pady=5)

        tk.Button(root, text='Запустить обработку', command=self.run_processing, bg='#4CAF50', fg='white',
                  font=('Segoe UI', 10, 'bold')).pack(padx=10, pady=15)

        self.log_text = scrolledtext.ScrolledText(root, wrap=tk.WORD, height=12, font=('Consolas', 9))
        self.log_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def select_inputFolder(self):
        folder = filedialog.askdirectory(title='Выберите папку со справками')
        if folder:
            self.inputFolder.set(folder)

    def select_journal(self):
        file = filedialog.askopenfilename(
            title='Выберите файл журнала',
            filetypes=[('Excel files', '*.xlsx')]
        )
        if file:
            self.journal.set(file)

    def run_processing(self):
        if not self.inputFolder.get():
            messagebox.showerror('Ошибка', 'Выберите папку со справками!')
            return
        if not self.journal.get():
            messagebox.showerror('Ошибка', 'Выберите файл журнала!')
            return

        startingTime = datetime.now()
        self.log_text.delete(1.0, tk.END)
        self.log(f'=== Начало обработки: {startingTime.strftime("%H:%M:%S")} ===')

        try:
            self.process_files()
            endingTime = datetime.now()
            timeDif = endingTime - startingTime
            self.log(f'\n=== Готово! Завершено в {endingTime.strftime("%H:%M:%S")}, время: {timeDif.total_seconds():.1f} сек ===')
            messagebox.showinfo('Успех', 'Обработка завершена!\nРезультат сохранён в файл с суффиксом "_new.xlsx"')
        except Exception as e:
            self.log(f'\n❌ КРИТИЧЕСКАЯ ОШИБКА: {e}')
            messagebox.showerror('Ошибка', f'Обработка прервана:\n{str(e)}')

    def process_files(self):
        folderPath = self.inputFolder.get()
        journalPath = self.journal.get()

        items = [f for f in os.listdir(folderPath) if not f.startswith('Журнал')]
        total_files = len(items)

        if total_files == 0:
            self.log("Нет файлов для обработки (файлы 'Журнал' игнорируются).")
            return

        Result = pd.DataFrame()
        failed_files = []
        empty_fields_warnings = []

        for idx, item in enumerate(items, start=1):
            filePath = os.path.join(folderPath, item)
            self.log(f"[{idx}/{total_files}] Обработка: {item}")

            try:
                if not filePath.endswith('.xls'):
                    try:
                        dataFile = pd.read_excel(filePath, engine='openpyxl').fillna('')
                    except Exception:
                        dataFile = pd.read_excel(filePath, engine='xlrd').fillna('')
                else:
                    dataFile = pd.read_excel(filePath, engine='xlrd').fillna('')

                # --- Дата и день недели ---
                original_date_cell = dataFile.iloc[1][2]
                converted_date = None
                weekDay = ''
                if original_date_cell != '' and not pd.isna(original_date_cell):
                    try:
                        converted_date = dateConvertion(original_date_cell)
                        if converted_date:
                            weekDay = rusLowerCaseWeekDays(converted_date.weekday())
                    except Exception:
                        weekDay = ''

                # --- ФИО водителя ---
                p = str(dataFile.iloc[18][6]).split(' ')
                shortName = f"{p[0]} {p[1][0].upper()+'.' if len(p) > 1 and p[1] else ''}{p[2][0].upper()+'.' if len(p) > 2 and p[2] else ''}".strip()

                # --- Пункт правил ---
                original_pdd_text = dataFile.iloc[40][1]
                pdd_value = ''
                if isinstance(original_pdd_text, str):
                    textLower = original_pdd_text.lower().strip()
                    hasPDD = 'пдд' in textLower
                    hasP = re.search(r'\bп\.?\s\d', textLower) is not None
                    if hasPDD or hasP:
                        pdd_value = pddChapter(original_pdd_text)

                # --- Стаж ---
                raw_stazh = str(dataFile.iloc[24][20]).strip()
                def process_stazh(raw):
                    if not raw or raw.lower() in ('', 'nan'):
                        return ''
                    if re.fullmatch(r'[\d\s.,]+', raw) and re.search(r'\d', raw):
                        num_part = re.search(r'\d+', raw)
                        if num_part:
                            return int(num_part.group())
                    return raw

                stazh_obshch = process_stazh(raw_stazh)
                stazh_v_parke = process_stazh(raw_stazh)

                # --- Пострадавшие ---
                postradavshie_raw = dataFile.iloc[14][8]
                postradavshie = 0
                if postradavshie_raw != '' and str(postradavshie_raw).lower() not in ('нет', ''):
                    try:
                        postradavshie = int(postradavshie_raw)
                    except:
                        postradavshie = 0

                # --- Координаты через Yandex ---
                addressText = dataFile.iloc[9][6]
                if pd.isna(addressText) or addressText == '':
                    lat, lon = '', ''
                else:
                    #normAddr = normalizeAddress(str(addressText))
                    lat, lon = yandex_geocode(addressText)
                    lat, lon = yandex_geocode2(addressText, YANDEX_API_KEY)
                    if lat == '' and lon == '':
                        self.log(f"⚠️ Не найдено: {addressText}")
                    time.sleep(0.1)  # Yandex: до 10 запросов/сек

                # --- DataFrame ---
                dfData = pd.DataFrame({
                    'Место': 1 if 'европротокол' in str(dataFile.iloc[66][7]).lower() else '',
                    'Дата ДТП': converted_date,
                    'Время ДТП': timeCheck(dataFile.iloc[1][17], item),
                    'День недели': weekDay,
                    ' Место ДТП (Адрес)': dataFile.iloc[9][6],
                    'Район': dataFile.iloc[10][9],
                    'Округ': dataFile.iloc[10][2],
                    'Координаты места ДТП (широта)': lat,
                    'Координаты места ДТП (долгота)': lon,
                    '3-я сторона': '',
                    'Название стороней организации': dataFile.iloc[74][6],
                    'Государственный регистрационный знак сторонего транспорта': dataFile.iloc[73][6],
                    'МГТ': '',
                    'Перевозчик': 'ГУП "Мосгортранс"',
                    'Название филиала': str(dataFile.iloc[17][4])
                        .replace('Филиал ', 'Ф').replace('Филилал ', 'Ф')
                        .replace('Южный', 'Ю').replace('Северный', 'С')
                        .replace('Юго-', 'Ю').replace('Северо-', 'С')
                        .replace('Восточный', 'В').replace('Западный', 'З'),
                    'Название площадки': dataFile.iloc[17][15],
                    'Маршрут': dataFile.iloc[31][8],
                    'Марка автобуса / электробуса': dataFile.iloc[32][8],
                    'Гаражный номер': '' if dataFile.iloc[33][20] == '' else int(dataFile.iloc[33][20]),
                    'Регистрационный номер': dataFile.iloc[33][8],
                    'Водитель': '',
                    'Табельный номер водителя': dataFile.iloc[19][6],
                    'ФИО водителя': shortName,
                    'Гражданство': dataFile.iloc[22][6],
                    'Возраст': dataFile.iloc[21][6],
                    'Стаж общий ': stazh_obshch,
                    'Стаж в парке': stazh_v_parke,
                    'ДТП': '',
                    'Вид ДТП': dataFile.iloc[7][3],
                    'Причина ДТП': dataFile.iloc[40][1],
                    'Виновник ДТП': str(dataFile.iloc[65][9])
                        .replace('Не вина', '3-е лицо')
                        .replace('Вина', 'Перевозчик')
                        .replace('В расследовании', 'Проводится разбор'),
                    'Пункт правил': pdd_value,
                    'Скорость по гланассу КМ/Ч': dataFile.iloc[36][16],
                    'Выделенная полоса (ДА; НЕТ,)': str(dataFile.iloc[12][9]).lower(),
                    'Пострадавшие': '',
                    'Кол-во пострадавших': postradavshie,
                    'в т.ч.    лёгкий   вред здоровью': postradavshie,
                    'в т.ч. средний вред здоровью': '0',
                    'в т.ч. тяжёлый вред здоровью': '0',
                    'Кол-во погибших': '0',
                    'ГК': '0',
                    'Ответст-сть': '',
                    'Постановление': dataFile.iloc[66][7],
                    'Дата постановления': (dateConvertion(dataFile.iloc[67][10])).strftime('%d.%m.%Y') if dateConvertion(dataFile.iloc[67][10]) else '',
                    'Наказание водителя': 'выговор' if dataFile.iloc[65][9] == 'Вина' else '',
                    'Проишествия': '',
                    'Резонансные проишествия': '',
                    'Проишествия с водителями': '',
                    'Проишествия с контролёрами': '',
                    'Проишествия с пассажирами': '',
                    'Кол-во задержек в движении': '',
                    'Сработка АНТИСОН': '',
                    'Кол-во отстранённых водителей.': '',
                    'Проишествия3': ''
                }, index=[0])

                # --- Проверка потерь данных ---
                fields_to_check = {
                    'Дата ДТП': original_date_cell,
                    'Время ДТП': dataFile.iloc[1][17],
                    'ФИО водителя': dataFile.iloc[18][6],
                    'Пункт правил': original_pdd_text,
                    'Стаж общий ': raw_stazh,
                    'Кол-во пострадавших': postradavshie_raw,
                }

                for col_name, orig_val in fields_to_check.items():
                    final_val = dfData[col_name].iloc[0]
                    if (
                        not pd.isna(orig_val) and str(orig_val).strip() != '' and
                        (pd.isna(final_val) or final_val == '')
                    ):
                        empty_fields_warnings.append(f"{item} → {col_name}: '{orig_val}' → ''")

                Result = pd.concat([Result, dfData], ignore_index=True)

            except Exception as e:
                failed_files.append(item)
                self.log(f"[{idx}/{total_files}] ❌ Ошибка: {e}")

        # --- Сортировка по дате ---
        if not Result.empty and 'Дата ДТП' in Result.columns:
            Result = Result.sort_values(by='Дата ДТП', na_position='last').reset_index(drop=True)

        # --- Сохранение в Excel ---
        output_path = journalPath[:-5] + '_new.xlsx'
        workbook = ox.load_workbook(journalPath)
        sheet = workbook['Лист1']

        filledColNums = [2, 3, 4, 5, 6, 7, 11, 12, 15, 16, 17, 18, 19, 20, 22, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 36, 43, 44]
        fillingBorder = Side(border_style='thin', color='AFC69F')
        fillingEmpty = PatternFill(fill_type='solid', start_color='ffffc000', end_color='ffffc000')

        for i, r in enumerate(dataframe_to_rows(Result, index=False, header=False)):
            clean_row = []
            for val in r:
                if val is None or (isinstance(val, float) and np.isnan(val)) or str(val).strip() == '':
                    clean_row.append('')
                else:
                    clean_row.append(val)

            sheet.append(clean_row)
            lastRowNew = sheet.max_row

            fillingStyle = PatternFill(
                fill_type='solid',
                start_color='ffc5e0b3' if i % 2 != 0 else 'ffe2efd9',
                end_color='ffc5e0b3' if i % 2 != 0 else 'ffe2efd9'
            )

            for cell in sheet[lastRowNew]:
                col_idx = cell.column
                val = cell.value
                if col_idx in (2, 43):
                    if isinstance(val, (datetime, date)) and not isinstance(val, str):
                        cell.number_format = 'DD.MM.YYYY'

                if cell.value == '' and col_idx in filledColNums:
                    cell.fill = fillingEmpty
                else:
                    cell.fill = fillingStyle

                cell.border = Border(left=fillingBorder, right=fillingBorder, top=fillingBorder, bottom=fillingBorder)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(output_path)
        self.log(f"\n✅ Результат сохранён: {output_path}")

        # --- Итоговый отчёт ---
        if failed_files:
            self.log("\n❌ Необработанные файлы:")
            for f in failed_files:
                self.log(f"  - {f}")
        else:
            self.log("\n✅ Все файлы обработаны.")

        if empty_fields_warnings:
            self.log(f"\n⚠️ Потери данных ({len(empty_fields_warnings)} записей):")
            for w in empty_fields_warnings[:50]:
                self.log(f"  - {w}")
            if len(empty_fields_warnings) > 50:
                self.log(f"  ... и ещё {len(empty_fields_warnings) - 50}")
        else:
            self.log("\n✅ Нет потерь данных.")


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()