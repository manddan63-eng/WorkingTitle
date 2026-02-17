
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import Label, filedialog, scrolledtext, messagebox
import threading
import numpy
import re
from datetime import datetime
#print(pd.__version__)
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import shutil
#import os
from openpyxl.utils import get_column_letter

contract_beg = datetime.now()

contract_end = datetime.now()

def is_row_empty(row):
    """
    Проверяет, является ли строка полностью пустой.
    Учитывает: None, пустые строки, пробелы, табуляции, неразрывные пробелы (\xa0).
    """
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str):
            # Удаляем ВСЕ виды пробельных символов (включая \xa0, \t, \n)
            cleaned = re.sub(r'[\s\u00A0]+', '', cell)
            if cleaned != '':
                return False
        else:
            # Любое не-строковое значение (число, дата, булево) считаем НЕПУСТЫМ
            return False
    return True

def clean_columns(df):
    new_columns = []
    rename_map = {}
    for i, col in enumerate(df.columns):
        # Определяем, является ли имя "пустым"
        is_empty_name = (pd.isna(col) or (isinstance(col, str) and 'Unnamed' in col.strip() ) )
        # Проверяем, есть ли непустые данные в столбце
        has_data = (not df[col].isnull().all() and not (df[col].astype(str).str.strip().replace('nan', '').eq('').all()))
        if is_empty_name:
            if has_data:
                # Столбец без имени, но с данными → переименовываем
                new_name = "VIN"
                new_columns.append(new_name)
                rename_map[col] = new_name
                #print(f"Переименован безымянный столбец в: {new_name}")
            else:
                # Столбец без имени и без данных → удаляем
                new_columns.append(None)  # помечаем на удаление
                #print(f"Удалён пустой безымянный столбец (индекс {i})")
        else:
            # Столбец с именем → оставляем как есть
            new_columns.append(col)
    # Удаляем помеченные столбцы
    cols_to_keep = [col for col in new_columns if col is not None]
    df = df[[orig_col for orig_col, keep in zip(df.columns, new_columns) if keep is not None]]
    # Применяем переименование
    if rename_map:
        # Обратите внимание: в df.columns сейчас старые имена (до фильтрации)
        # Поэтому лучше переименовать после отбора нужных столбцов
        final_rename = {}
        current_cols = list(df.columns)
        for old_name, new_name in rename_map.items():
            if old_name in current_cols:
                final_rename[old_name] = new_name
        if final_rename:
            df = df.rename(columns=final_rename)
    return df
def extract_period_from_merged_cells(text):
    """Ищет текст 'за период ... г. по ... г.' в объединённых ячейках первых 15 строк"""
    # Ищем обе возможные формы: с точками и с месяцами
    if 'за период' in text.lower() and 'по' in text.lower() and 'г.' in text:
        # Формат 1: 01.01.2024
        dates_dot = re.findall(r'\b\d{1,2}\.\d{1,2}\.\d{4}\b', text)
        # Формат 2: 1 января 2024
        dates_word = re.findall(r'\b(\d{1,2})\s+(' + '|'.join(RU_MONTHS.keys()) + r')\s+(\d{4})\b',text,re.IGNORECASE)
        if len(dates_dot)>2:
            dates_dot = dates_dot[1:3]
        if len(dates_word)>2:
            dates_word = dates_word[1:3]
        dates_parsed = []
        # Обрабатываем даты с точками
        if not dates_word:
            for d in dates_dot:
                #print('dots')
                #print(dates_dot)
                try:
                    dt =  pd.to_datetime(d, format = '%d.%m.%Y')
                    dates_parsed.append(dt)
                except ValueError as e:
                    #print(e)
                    pass
        else:
                # Обрабатываем даты со словами
            for day, month_word, year in dates_word:
                month_num = RU_MONTHS.get(month_word.lower())
                #print(month_num)
                #print(dates_word)
                if month_num:
                    try:
                        dt_str = f"{int(day):02d}.{month_num}.{year}"
                        #print(dt_str)
                        dt = pd.to_datetime(dt_str, format= '%d.%m.%Y')
                        #print('asd')
                        #print(dt)
                        dates_parsed.append(dt)
                        #print(dates_parsed)
                    except ValueError as e:
                        #print(e)
                        pass
        # Сортируем по времени (на случай, если порядок нарушен)
        dates_parsed = sorted(dates_parsed)
    
        if dates_parsed[0]>=dates_parsed[-1]:
            period_start = dates_parsed[-1]
            period_end = dates_parsed[0]
        else:
            period_start = dates_parsed[0]
            period_end = dates_parsed[-1] 
        period_display = f"за период {period_start.strftime('%d.%m.%Y')} г. по {period_end.strftime('%d.%m.%Y')} г."
        #contract_beg, contract_end = period_start, period_end
        #period_extracted = True
        #log_to_gui(f"Период из отчёта (openpyxl): {period_display}")
        return period_start, period_end, period_display
    return None, None, None
def normalize_date(val):
    """
    Превращает любое значение даты в строку 'ДД.ММ.ГГГГ' или оставляет пустым.
    Поддерживает: datetime, строку 'ДД.ММ.ГГГГ', число Excel, float, int.
    """
    if pd.isna(val) or val in ('', 'nan', None):
        return ''
    # Случай 1: уже строка в правильном формате
    if isinstance(val, str):
        val = val.strip()
        # Проверим, похоже ли на ДД.ММ.ГГГГ
        if re.fullmatch(r'\d{1,2}\.\d{1,2}\.\d{4}', val):
            try:
                dt = datetime.strptime(val, '%d.%m.%Y')
                return dt.strftime('%d.%m.%Y')
            except ValueError:
                pass
        # Если строка — попробуем распарсить гибко
        try:
            dt = pd.to_datetime(val, errors='coerce', dayfirst=True)
            if pd.notna(dt):
                return dt.strftime('%d.%m.%Y')
        except Exception:
            return val  # оставляем как есть, если не удалось    
    # Случай 2: число (возможно, Excel serial date)
    if isinstance(val, (int, float)):
        if val <= 0:
            return ''
        # Попробуем как Excel-дату (1 = 01.01.1900)
        try:
            # Excel считает с 1900-01-01, но имеет баг с 1900 (високосный)
            # Используем pd.to_datetime с origin='1899-12-30'
            dt = pd.to_datetime(val, unit='D', origin='1899-12-30', errors='coerce')
            if pd.notna(dt):
                return dt.strftime('%d.%m.%Y')
        except Exception:
            pass        
    # Случай 3: datetime
    if hasattr(val, 'strftime'):
        return val.strftime('%d.%m.%Y')    
    # По умолчанию — вернуть строку
    return str(val).strip()
def extract_number_from_result(val):
    if pd.isna(val) or val == '':
        return 0
    #if 'стажер' in val:
        #return 0
    if val<=0:
        return 0
    # Ищем первое число в строке (может быть отрицательным)
    else:
        return val
def read_excel_as_strings(file_path, sheet_name=None):
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True, read_only=True)
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(values_only=True):
        # Проверка на полностью пустую строку
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            if data: # если уже есть данные — останавливаемся
                break
            else:     # если ещё нет данных — пропускаем
                continue
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append('')
            elif isinstance(cell, str):
                cleaned_row.append(cell.strip())
            elif isinstance(cell, (int, float)):
                # Если целое число (в т.ч. 123.0) → '123'
                if isinstance(cell, float) and cell.is_integer():
                    cleaned_row.append(str(int(cell)))
                else:
                    # Не целое — оставляем как есть (редко для гаражных номеров)
                    cleaned_row.append(str(cell))
            elif hasattr(cell, 'strftime'):  # datetime
                # Форматируем дату как ДД.ММ.ГГГГ
                cleaned_row.append(cell.strftime('%d.%m.%Y'))
            else:
                cleaned_row.append(str(cell))
        data.append(cleaned_row)
    wb.close()
    if not data:
        return pd.DataFrame()
    headers = data[0]
    rows = data[1:]
    # Выравнивание колонок (на случай разного числа ячеек)
    max_len = max(len(headers), *(len(r) for r in rows)) if rows else len(headers)
    headers = (headers + [''] * max_len)[:max_len]
    rows = [(r + [''] * max_len)[:max_len] for r in rows]
    return pd.DataFrame(rows, columns=headers)

RU_MONTHS = {'января': '01','февраля': '02','марта': '03','апреля': '04','мая': '05','июня': '06','июля': '07','августа': '08','сентября': '09','октября': '10','ноября': '11','декабря': '12'}
short_names = {'ФСВ':'Северо-Восточный','ФСЗ':'Северо-Западный','ФЮ':'Южный','СВ':'Северо-Восточный','СЗ':'Северо-Западный', 'Северо-восточный':'Северо-Восточный','Северо-западный':'Северо-Западный', 'северо-восточный':'Северо-Восточный','северо-западный':'Северо-Западный','южный':'Южный'}

def get_last_row_with_data(worksheet, col):
    for row in range(worksheet.max_row, 0, -1):
        if worksheet.cell(row=row, column=col).value not in (None, "", " "):
            return row
    return 1

def log_to_gui(message):
    if 'log_text' in globals() and log_text:
        log_text.configure(state='normal')
        log_text.insert(tk.END, message + "\n")
        log_text.configure(state='disabled')
        log_text.yview(tk.END)
    else:
        print(message) 

# Глобальная переменная для хранения пути к файлу отчета
global_report_file = None

def start_gui():
    global log_text, global_report_file
    global_report_file = None  # Сброс при запуске
    
    root = tk.Tk()
    root.configure(background='lightblue', border = (10) )
    #root.BackgroundColor('blue')
    root.title("Сравнение данных МГТ с вигитон/антисон")
    root.geometry("500x480")
    log_text = scrolledtext.ScrolledText(root, state='disabled', wrap=tk.WORD, font=("Moscow Sans", 10))
    
    # Кнопка выбора файла отчета (вигитон/антисон)
    def on_select_report_file():
        nonlocal root
        file_path = filedialog.askopenfilename(
            title="Выберите файл отчета (вигитон/антисон)",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_path:
            global global_report_file
            global_report_file = Path(file_path)
            log_to_gui(f"Выбран файл отчета: {global_report_file.name}")
        else:
            log_to_gui("Файл отчета не выбран.")
    tk.Label(root, text='В данной программе в обработку берутся файлы:\n- "Отчет Вигитон.xlsx" или "Отчет Антисон.xlsx" - выбирается 1 файл\n- файл(ы) МГТ о ТС формата ".xlsx" - можно выбрать несколько', font=('Moscow Sans', 9)).pack(anchor='w', padx=10, pady=(10, 0), fill = tk.X)
    tk.Label(root, text='Нажав на кнопки ниже - откроется проводник\nдля выбора файла(ов).\n После выбора файла(ов) нажать на кнопку "Открыть" в правом нижнем углу', font=('Moscow Sans', 9)).pack(anchor='w', padx=15, pady=(15, 0), fill = tk.X)
    tk.Label(root, text = 'УКАЗАННЫЕ ФАЙЛЫ НЕ ДОЛЖНЫ БЫТЬ ОТКРЫТЫ НА МОМЕНТ ЧТЕНИЯ', font=('Moscow Sans', 9)).pack(padx= 15, pady=(15, 0), fill = tk.X)
    btn_report = tk.Button(root, text="Выбрать файл отчета (вигитон/антисон)", command=on_select_report_file, font=("Moscow Sans", 11), padx=15, pady=8, bg='#E0F7FA')
    btn_report.pack(pady=5)
    
    # Метка-разделитель
    Label(root, text="↓", font=("Moscow Sans", 14), bg='lightblue').pack(pady=2)
    
    # Кнопка выбора контрольных файлов
    def on_select_files():
        nonlocal root
        file_paths = filedialog.askopenfilenames(
            title="Выберите контрольные Excel-файлы (.xlsx)",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_paths:
            # Передаем оба параметра: контрольные файлы + файл отчета
            threading.Thread(target=lambda: process_files(file_paths, global_report_file), daemon=True).start()
        else:
            log_to_gui("Контрольные файлы не выбраны.")
    
    btn_files = tk.Button(root, text="Выбрать контрольные файлы", command=on_select_files, font=("Moscow Sans", 12), padx=20, pady=10, bg='#B3E5FC')
    btn_files.pack(pady=8)
    
    log_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
    root.mainloop()

def process_files(file_paths, report_file_path=None):
    global log_text
    input_file_paths = []
    contract_name = ''
    combined_df = pd.DataFrame()
    
    # Обработка контрольных файлов
    for fp in file_paths:
        filename = Path(fp).name
        if filename.lower().endswith('.xlsx'):
            input_file_paths.append(Path(fp))
        else:
            log_to_gui(f"Пропущен файл: {filename}")
    
    # Обработка файла отчета (если выбран)
    if report_file_path and report_file_path.exists():
        contract_path = report_file_path
        contract_name = report_file_path.name
        log_to_gui(f"Файл отчета для сравнения: {contract_name}")
    else:
        log_to_gui("Файл отчета (вигитон/антисон) не выбран. Будет выполнена только обработка контрольных файлов.")
    
    if not input_file_paths:
        messagebox.showerror("Ошибка", "Не найдено подходящих контрольных файлов")
        return
    log_to_gui(f"\nНайдено контрольных файлов: {len(input_file_paths)}")
    startingTime = datetime.now()
    log_to_gui(f'=== Начало чтения: {startingTime.strftime("%H:%M:%S")} ===')
    for file_path in input_file_paths:
        filename = file_path.name    
        period_display = ""
        success = False
        df = pd.DataFrame()
        # Попытка 1: openpyxl
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
            ws = wb.active
            filial = ws.cell(row = 3, column = 3).value
            if 'Организация: ' in filial:
                filial = filial.split('Организация: ')[1]
            period_display =  ws.cell(row = 2, column = 3).value
            #ws.unmerge_cells(start_row=1, start_column=2, end_row=2, end_column=5)
            if filial:
                for key, value in short_names.items():
                    filial = filial.replace(key, value)
                log_to_gui(f"В файле с данными о филиале {filial}")
            if period_display:
                log_to_gui(f"{period_display}")
            #ws.unmerge_cells(start_row=6, start_column=1, end_row=6, end_column=12)
            #ws.unmerge_cells(start_row=6, start_column=13, end_row=7, end_column=13)
            #ws.unmerge_cells(start_row=6, start_column=14, end_row=7, end_column=14)
            #ws.cell(row = 7, column = 13, value = ws.cell(row = 6, column = 13).value)
            #ws.cell(row = 7, column = 14, value = ws.cell(row = 6, column = 14).value)
            header_row = 0
            for row_idx, row in enumerate(ws.iter_rows(), start=header_row-1):
                row_data = [cell.value for cell in row]
                if 'Дата' in row_data:
                    header_row = row_idx
                    break
            #print(header_row)
            # Коррекция: реальный номер строки = header_row + 2 (из-за start=header_row-1 при header_row=0)
            real_header_row = header_row + 2
            
            # Читаем строку заголовка и определяем столбцы БЕЗ None (столбцы с объединенными ячейками имеют None)
            header_cells = ws[real_header_row]
            keep_columns = [i for i, cell in enumerate(header_cells) if cell.value is not None]
            headers = [header_cells[i].value for i in keep_columns]
            
            # Собираем данные: только строки с точкой в первом столбце (даты в формате ДД.ММ.ГГГГ)
            data = []
            row_counter = 0
            for row in ws.iter_rows(min_row=real_header_row+1, values_only=True):
                # Прерываем чтение при первой полностью пустой строке
                if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                    break
                # Фильтр: только строки с точкой в первом столбце (дата)
                if row and len(row)>0 and isinstance(row[0], str): 
                    dates_word = re.findall(r'\b\d{1,2}\.\d{1,2}\.\d{4}\b',row[0],re.IGNORECASE)
                    #if row_counter>2070:
                        #print('===')
                        #print(row_counter)
                        #print(dates_word)
                        #print('==')
                    if '.' in row[0] and len(dates_word)==1:
                        #print('+')
                        # Берем только столбцы без None в заголовке
                        filtered_row = [row[i] if i < len(row) else None for i in keep_columns]
                        data.append(filtered_row)
                        row_counter += 1
            
            # Создаем DataFrame
            if data and headers:
                # Выравнивание строк по количеству заголовков
                max_cols = len(headers)
                aligned_data = []
                for row in data:
                    if len(row) < max_cols:
                        row += [None] * (max_cols - len(row))
                    aligned_data.append(row[:max_cols])
                df = pd.DataFrame(aligned_data, columns=headers)
                log_to_gui(f"Прочитано {row_counter} строк с датами из файла {filename}")

                col_map = {}
                for col in df.columns:
                    col_str = str(col).strip().lower()
                    if 'часы' in col_str or 'длительн' in col_str :
                        col_map[col] = 'Часы'
                    if 'гар.' in col_str or  'гаражн' in col_str:
                        col_map[col] = 'Гаражный номер ТС'
                    if 'vin' in col_str:
                        col_map[col] = 'VIN'
                    if 'пробег' in col_str or 'длин' in col_str:
                        col_map[col] = 'Пробег'
                    if 'дата' == col_str:
                        col_map[col] = 'Дата'
                    if 'номер' == col_str:
                        col_map[col] = 'Номер'
                df = df.rename(columns=col_map)

                # Приведение ключевых колонок к строке сразу после переименования
                for col in ['Гаражный номер ТС', 'Номер', 'VIN']:
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.strip().replace('nan', '')
                #required = ['Дата', 'Гаражный номер ТС', 'Номер', 'VIN', 'Часы', 'Пробег']
                #if all(c in df.columns for c in required[:3]):
                df['Часы'] = pd.to_numeric(df['Часы'], errors='coerce')
                df['Пробег'] = pd.to_numeric(df['Пробег'], errors='coerce')
                df= df.groupby(['Дата','Гаражный номер ТС','VIN' ]).agg({'Часы':'sum', 'Пробег':'sum'})
                df = df[(df['Часы'] >= 2) & (df['Пробег'] >= 20)].reset_index()[['Дата', 'Гаражный номер ТС', 'VIN']]
                #log_to_gui(f"По количеству часов и пробегу в файле {filename} есть {len(df)} строк")
                #df = df[required[:3]].copy()
                if 'Дата' in df.columns:
                    df['Дата'] = df['Дата'].apply(normalize_date)

                #result = df.groupby(['Гаражный номер ТС', 'VIN']).size().reset_index(name='Количество по МГТ')

                #df["Количество по МГТ"] = df.groupby(['Гаражный номер ТС', 'VIN']).size()
                #df["Количество по МГТ"] = df.groupby(['Гаражный номер ТС', 'VIN'])['Дата'].transform('nunique')
                df = df.groupby(['Гаражный номер ТС', 'VIN']).size().reset_index(name='Количество по МГТ').astype(str)
                df['Филиал'] = filial

                # Альтернатива (ещё короче):
                # df["Количество по МГТ"] = df.groupby(['Гаражный номер ТС', 'VIN']).transform('size')
                log_to_gui(f"В файле {filename} есть информация о количестве дней использования {len(df)} ТС")

            else:
                df = pd.DataFrame()
                
                log_to_gui(f"Не найдено строк с датами в файле {filename}")

            wb.close()
            success = True
        except Exception as e2:
            log_to_gui(f"openpyxl не смог прочитать '{filename}': {e2}")
            success = False
        except Exception as  e:
            print(e)
        combined_df = pd.concat([combined_df, df], ignore_index=True)

    if len(combined_df)>0:
        log_to_gui('Данные из файлов из МГТ собраны')
    #print(combined_df)
    used_filials = combined_df['Филиал'].unique().tolist()
    #print(used_filials)
    if contract_path:
        contract_dataframes = pd.DataFrame()
        period_extracted = False
        # Попытка 1: pandas (редко работает для многолистовых, но попробуем)
        try:
            xls = pd.ExcelFile(contract_path)
            sheetNames = xls.sheet_names
            for sheet in sheetNames:
                if (sheet in used_filials):
                    df_temp = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)
                    # Ищем строку с "за период"
                    for idx, row in df_temp.iterrows():
                        cell_val = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
                        if 'за период' in cell_val.lower() and 'по' in cell_val.lower() and 'г.' in cell_val:
                            period_start, period_end, period_display = extract_period_from_merged_cells(cell_val.lower())
                            contract_beg, contract_end = period_start, period_end
                            period_extracted=True



                            df_sheet = pd.read_excel(xls, sheet_name=sheet, skiprows=1, dtype=str)
                            #print(sheet)
                            if 'Гаражный номер ТС' in df_sheet.columns:
                                # ← ДОБАВИТЬ ЭТОТ БЛОК:
                                # Остановка на первой полностью ппустой строке
                                #mask_empty = df_sheet.apply(lambda row: row.astype(str).str.strip().eq('').all(), axis=1)
                                mask_empty = df_sheet.apply(lambda row: is_row_empty(row.values), axis=1)
                                first_empty_idx = mask_empty.idxmax() if mask_empty.any() else len(df_sheet)
                                df_sheet = df_sheet.iloc[:first_empty_idx].copy()
            
                                            # === КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: ОБРЕЗКА ПО ПЕРВОЙ ПУСТОЙ СТРОКЕ ===
            # Используем надёжную функцию is_row_empty

                                #df_sheet = df_sheet[[ 'Количество ед.','Гаражный номер ТС', 'VIN']].copy()
                                df_sheet['Филиал'] = sheet
                                df_sheet = clean_columns(df_sheet)
                                contract_dataframes= pd.concat([contract_dataframes,df_sheet],ignore_index =False)

                            for i in range(0, 10):
                                if df_sheet.iloc[i, 0]=='№':
                                    header_row = i
                                    df_sheet = pd.read_excel(xls, sheet_name=sheet, skiprows=2+i, dtype=str)
                            if 'Гаражный номер ТС' in df_sheet.columns :
                                #df_sheet = df_sheet[[ 'Количество ед.','Гаражный номер ТС', 'VIN']].copy()
                                df_sheet['Филиал'] = sheet
                                df_sheet = clean_columns(df_sheet)
                                contract_dataframes= pd.concat([contract_dataframes,df_sheet],ignore_index =False)
            if period_extracted:
                log_to_gui('В отчете указаны данные '+period_display)
            xls.close()
            #log_to_gui(f"pandas смог прочитать отчёт: {period_display}")
        except Exception as e:
            log_to_gui(f"pandas не смог прочитать отчёт: {e}")
        # Попытка 2: openpyxl (основной метод для отчёта)
        if not period_extracted:
            try:
                wb = openpyxl.load_workbook(contract_path, data_only=True, read_only=True)
                for sheet_name in wb.sheetnames:
                    if (sheet in used_filials):
                        ws = wb[sheet_name]
                        for row_idx in range(1, min(16, ws.max_row + 1)):
                            cell_val = ws.cell(row=row_idx, column=1).value
                            if cell_val and isinstance(cell_val, str):
                                text = cell_val.strip()
                                if 'за период' in text.lower() and 'по' in text.lower() and 'г.' in text:
                                    period_start, period_end, period_display = extract_period_from_merged_cells(text.lower())
                                    contract_beg, contract_end = period_start, period_end
                                    period_extracted=True

                            #if is_row_empty(row)== True:
                                #break


                        header_row = None
                        headers = []
                        for r in range(1, min(11, ws.max_row + 1)):
                            row_vals = [str(ws.cell(row=r, column=c).value or '').strip() for c in range(1, 6)]
                            if 'Гаражный номер ТС' in ''.join(row_vals) :
                                header_row = r
                                headers = row_vals
                                break
                        if header_row:
                            rows = []
                            for r in range(header_row + 1, ws.max_row + 1):

                                vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                                if all(v is None or str(v).strip() == '' for v in vals):
                                    break                                
                                if is_row_empty(row)== True:
                                    print('Сработала функция')
                                    break
                                if vals and  len(vals)>0 and vals[0]=='':
                                    print('Сработало условие')
                                    break
                                if r>2070:
                                    print('----')
                                    print(r)
                                    print(vals)
                                    print('====')                                
                                if not isinstance(vals[0], (int)):
                                    break


                                rows.append(vals)
                            if rows:                                

                                df_sheet = pd.DataFrame(rows, columns=headers)
                                cols_map = {}
                                for col in df_sheet.columns:
                                    if 'количество' in col.lower():
                                        cols_map[col] = 'Количество ед.'
                                    elif ('гар.' in col.lower() and '№' in col.lower()) or ('гар' in col.lower() and 'ТС' in col.lower()):
                                        cols_map[col] = 'Гаражный номер ТС'
                                    elif 'vin' in col.lower():
                                        cols_map[col] = 'VIN'
                                if cols_map:
                                    df_sheet = df_sheet[list(cols_map.keys())].rename(columns=cols_map)
                                    df_sheet['Филиал'] = sheet_name
                                    df_sheet = clean_columns(df_sheet)
                                    print(df_sheet['Количество'].dtype)
                                    #df_sheet = df_sheet[df_sheet['Количество']!= 0 and df_sheet['Количество']!= '' and df_sheet['Количество']!= '0']
                                    contract_dataframes= pd.concat([contract_dataframes,df_sheet],ignore_index =False)
                wb.close()
                if period_extracted:
                    log_to_gui('В отчете указаны данные '+period_display)
            except Exception as e:
                log_to_gui(f"openpyxl не смог прочитать отчёт: {e}")

        if len(contract_dataframes)==0:
            log_to_gui("Данные из отчёта не извлечены")
        #print(contract_dataframes.columns.to_list())
    
    log_to_gui("Идет подсчет разницы в количестве с МГТ...")
    # Подготовка данных

    if len(contract_dataframes)!=0:
        combined_df['Филиал'] = (combined_df['Филиал'].astype(str).str.strip().replace({'nan': ''}))
        combined_df['Филиал'] = combined_df['Филиал'].map(short_names).fillna('Неизвестно')
        combined_df['Гаражный номер ТС'] = (pd.to_numeric(combined_df['Гаражный номер ТС'], errors='coerce').fillna(0).astype('Int64').astype(str).str.zfill(6).replace('000000', ''))
        #used_filials = combined_df['Филиал'].unique().tolist()
        #contract_dataframes = contract_dataframes[contract_dataframes['Филиал'] in used_filials]
            # Фильтр по периоду
        #if contract_beg and contract_end:
            #combined_df['Дата'] = pd.to_datetime(combined_df['Дата'], format='%d.%m.%Y', errors='coerce')
            #combined_df = combined_df[(combined_df['Дата'] >= period_start) & (combined_df['Дата'] <= period_end)]
            #combined_df['Дата'] = combined_df['Дата'].dt.strftime('%d.%m.%Y')
        contract_dataframes['Гаражный номер ТС'] = (pd.to_numeric(contract_dataframes['Гаражный номер ТС'], errors='coerce').fillna(0).astype('Int64').astype(str).str.zfill(6).replace('000000', ''))
        contract_dataframes = contract_dataframes.iloc[( contract_dataframes['VIN']!='') & (contract_dataframes['Количество ед.']!= 0) & (contract_dataframes['Количество ед.']!= '0')]
        checking = pd.merge(contract_dataframes, combined_df, left_on= ['Филиал', 'VIN', 'Гаражный номер ТС'], right_on=['Филиал','VIN','Гаражный номер ТС'], how = 'left', suffixes=['', 'CUB'])
        checking = checking[['№','Наименование услуги','Единица измерения','Количество ед.','Гаражный номер ТС','Государственный номер ТС','Цена 1 ед., руб.','Итого, руб.','Количество по МГТ','VIN', 'Филиал']]
        checking['Количество по МГТ'] = checking['Количество по МГТ'].fillna(0)
        checking['Разница в количестве с МГТ'] = (pd.to_numeric(checking['Количество ед.'], errors='coerce').fillna(0).astype(int) - pd.to_numeric(checking['Количество по МГТ'], errors='coerce').fillna(0).astype(int))
        #checking['Разница в количестве с МГТ'] = (str(checking['Разница в количестве с МГТ'])+', стажер') if checking['Дата выгрузки']=='' else (str(checking['Разница в количестве с МГТ']))
        log_to_gui("Идет подсчет разницы по сумме с МГТ...")
        # Извлекаем число из "Разница в количестве с МГТ" (игнорируем ", стажер") и умножаем на цену
        checking['Разница по сумме с МГТ'] = (checking['Разница в количестве с МГТ'].apply(extract_number_from_result) * pd.to_numeric(checking['Цена 1 ед., руб.'], errors='coerce').fillna(0)).round(2)
        #print(checking['Разница по сумме с МГТ'].dtype)
        diif_sum = checking['Разница по сумме с МГТ'].sum()
        log_to_gui("Разница в количестве с МГТ и разница по сумме подсчитаны")
        #  Разница в количестве с МГТ
        with pd.ExcelWriter('Свод_по_собранным_данным.xlsx') as writer:
            checking.to_excel(writer, index=False, sheet_name='Свод')
            #practice.to_excel(writer, index=False, sheet_name='стажировки')
            #combined_df.to_excel(writer, index=False, sheet_name='КУБы')
            #contract_dataframes.to_excel(writer, index=False, sheet_name='Контракты')
        log_to_gui(f"\nСоздан файл 'Свод_по_собранным_данным.xlsx', в котором хранится сводная таблица по данным из отчета, данных МГТ")

        if contract_path:
            # Создаём копию отчёта
            new_file_name = 'Копия_отчета.xlsx'
            shutil.copyfile(contract_path, new_file_name)
            log_to_gui(f"\nСоздана копия отчёта '{contract_name}': {new_file_name}.\nВ этот файл будут внесены разница в количестве с МГТ, разница в сумме с МГТ:")
            # Открываем копию через openpyxl для сохранения стилей (БЕЗ read_only!)
            wb = openpyxl.load_workbook(new_file_name)
            # Филиалы для обработки (только нужные три)
            #target_filials = combined_df['Филиал'].unique().tolist()
            target_filials = checking['Филиал'].unique().tolist()
            processed_filials = []
            for sheet_name in wb.sheetnames:
                # Проверяем, является ли лист нужным филиалом (регистронезависимо)
                if any(tf.lower() in sheet_name.lower() for tf in target_filials):
                    ws = wb[sheet_name]
                    filial_name = next(tf for tf in target_filials if tf.lower() in sheet_name.lower())
                    processed_filials.append(filial_name)
                    # Фильтруем данные для этого филиала
                    filial_data = checking[checking['Филиал'] == filial_name].copy()
                    if filial_data.empty:
                        log_to_gui(f"  >> Филиал '{filial_name}' (лист '{sheet_name}'): нет данных для записи")
                        continue
                    # Определяем строку с заголовками (ищем "№" в первых 10 строках)
                    header_row = None
                    col_index_number = None  # индекс столбца "№"
                    col_index_number = None  # индекс столбца "№"
                    for r in range(1, min(11, ws.max_row + 1)):
                        for c in range(1, min(10, ws.max_column + 1)):
                            cell_val = str(ws.cell(row=r, column=c).value or '').strip()
                            if cell_val == '№':
                                header_row = r
                                col_index_number = c
                                break
                        if header_row:
                            break
                    if not header_row:
                        log_to_gui(f"   Не найден заголовок '№' в листе '{sheet_name}'")
                        continue
                    # Собираем маппинг: номер услуги → номер строки в Excel
                    number_to_row = {}
                    for r in range(header_row + 1, ws.max_row + 1):
                        cell_val = str(ws.cell(row=r, column=col_index_number).value or '').strip()
                        if cell_val and cell_val != '':
                            number_to_row[cell_val] = r
                    # Записываем данные для каждой найденной услуги
                    written_count = 0
                    for _, row in filial_data.iterrows():
                        service_num = str(row['№']).strip()
                        result_val = str(row['Разница в количестве с МГТ'])
                        diff_val = row['Разница по сумме с МГТ']
                        if service_num in number_to_row:
                            target_row = number_to_row[service_num]
                            # Находим последний непустой столбец в строке
                            last_col = ws.max_column
                            while last_col >= 1 and ws.cell(row=target_row, column=last_col).value in (None, '', ' '):
                                last_col -= 1
                            ws.cell(row=header_row, column=last_col + 1, value='Разница в количестве с МГТ')
                            ws.cell(row=header_row, column=last_col + 2, value='Разница по сумме с МГТ')
                            # Определяем столбцы для записи с обходом объединённых ячеек
                            # Начинаем со столбца после последнего заполненного
                            result_col = last_col + 1
                            # Сдвигаемся вправо, пока не найдём обычную ячейку (не MergedCell)
                            while result_col <= ws.max_column + 10:  # +10 на случай длинных объединений
                                cell = ws.cell(row=target_row, column=result_col)
                                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                                    break
                                result_col += 1
                            diff_col = result_col + 1
                            while diff_col <= ws.max_column + 10:
                                cell = ws.cell(row=target_row, column=diff_col)
                                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                                    break
                                diff_col += 1
                            # Записываем значения в найденные обычные ячейки
                            ws.cell(row=target_row, column=result_col, value=result_val)
                            ws.cell(row=target_row, column=diff_col, value=diff_val)
                            # Применяем базовое форматирование (как в ваших требованиях)
                            for col in [result_col, diff_col]:
                                cell = ws.cell(row=target_row, column=col)
                                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                                cell.font = openpyxl.styles.Font(name='Moscow Sans', size=10)
                        
                            written_count += 1
                    #print(diff_col)
                    #print(header_row+written_count)
                    ws.cell(row = written_count-2, column = ws.max_column, value = diif_sum)

                    log_to_gui(f"  >> Филиал '{filial_name}' (лист '{sheet_name}'): записано {written_count} строк")
            # Сохраняем изменения
            wb.save(new_file_name)
            wb.close()

            log_to_gui(f"  Обработаны филиалы: {', '.join(processed_filials) if processed_filials else 'нет подходящих листов'}")
        else:
            log_to_gui("Файл отчёта не был загружен — запись в копию пропущена")

    endingTime = datetime.now()
    timeDif = endingTime - startingTime

    log_to_gui(f'\n=== Времени потрачено на чтение : {timeDif.total_seconds():.1f} сек ===')
    log_to_gui("Обработка завершена")
    messagebox.showinfo('Успех', 'Обработка завершена!\nРезультаты сохранены.')

if __name__ == "__main__":
    start_gui()